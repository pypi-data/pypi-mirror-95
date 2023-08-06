# -*- coding: utf-8 -*-

import xlrd
import xlwt
from enum import Enum
from xlutils import copy
from openpyxl import Workbook


class CellDataType(Enum):
    empty = 0
    string = 1
    number = 2
    date = 3
    boolean = 4
    error = 5
    blank = 6


class WriteExcel(object):

    def create_file(self, is_xlsx=False):
        if is_xlsx:
            self.book_wt_obj = Workbook()
        else:
            self.book_wt_obj = xlwt.Workbook(encoding="utf-8",
                                             style_compression=0)

        self.sheet_list = []
        self.is_xlsx = is_xlsx
        self.is_frist_sheet = True

    def add_sheet(self, sheet_name):
        if self.is_xlsx:
            if self.is_frist_sheet:
                self.is_frist_sheet = False
                self.sheet_wt_obj = self.book_wt_obj.active
                self.sheet_wt_obj.title = sheet_name
            else:
                self.sheet_wt_obj = self.book_wt_obj.create_sheet(sheet_name)

        else:
            self.sheet_wt_obj = self.book_wt_obj.add_sheet(sheet_name,
                                                           cell_overwrite_ok=True)

        self.sheet_list.append(self.sheet_wt_obj)

    def set_active_sheet(self, sheet_index):
        self.sheet_wt_obj = self.sheet_list[sheet_index - 1]

    def set_cell_data(self, row_index, column_index, data):
        if self.is_xlsx:
            '''    
            self.sheet_wt_obj.cell(row_index,
                                   column_index).value = data
            '''
            self.sheet_wt_obj.cell(row=row_index, column=column_index, value=data)
        else:
            self.sheet_wt_obj.write(row_index - 1, column_index - 1, data)

    def save_as(self, file_name):
        if self.is_xlsx:
            self.book_wt_obj.save(file_name)
        else:
            if file_name.endswith(".xlsx"):
                raise RuntimeError("cannot save file as .xlsx,please use ."
                                   "xls In this mode")
            self.book_wt_obj.save(file_name)


class ReadExcel(object):

    book_read_obj = None
    sheet_read_obj = None
    book_write_obj = None
    sheet_write_obj = None
    file_name = ""

    def open_file(self, file_name):
        self.book_read_obj = xlrd.open_workbook(file_name)
        self.book_write_obj = copy.copy(self.book_read_obj)
        self.file_name = file_name

    def set_active_sheet(self, sheet_index):
        sheet_index = sheet_index - 1
        self.sheet_read_obj = self.book_read_obj.sheet_by_index(sheet_index)
        self.sheet_write_obj = self.book_write_obj.get_sheet(sheet_index)

    def __set_active_sheet_by_name(self, sheet_name):
        self.sheet_read_obj = self.book_read_obj.sheet_by_name(sheet_name)

    def get_sheet_name(self, sheet_index):
        return self.book_read_obj.sheet_names()[sheet_index - 1]

    def get_sheet_names(self):
        return self.book_read_obj.sheet_names()

    def get_cell_data(self, row_index, column_index):
        return self.sheet_read_obj.cell_value(row_index - 1, column_index - 1)

    def get_row_data(self, row_index):
        return self.sheet_read_obj.row_values(row_index - 1)

    def get_column_data(self, column_index):
        return self.sheet_read_obj.col_values(column_index - 1)

    def get_row_count(self):
        return self.sheet_read_obj.nrows

    def get_column_count(self):
        return self.sheet_read_obj.ncols

    def get_cell_type(self, row_index, column_index):
        return self.sheet_read_obj.cell(row_index - 1, column_index - 1).ctype

    def get_multiple_row_data(self, begin_row_index=1):
        begin_row_index = begin_row_index - 1
        for row_index in range(begin_row_index, self.sheet_read_obj.nrows):
            yield self.sheet_read_obj.row_values(row_index)

    def get_multiple_column_data(self, begin_column_index=1):
        begin_column_index = begin_column_index - 1
        for column in range(begin_column_index, self.sheet_read_obj.ncols):
            yield self.sheet_read_obj.col_values(column)

    def set_cell_data(self, row_index, column_index, data):
        return self.sheet_write_obj.write(row_index - 1,
                                          column_index - 1, data)

    def save_as(self, file_name=""):
        if file_name == "":
            file_name = self.file_name

        if file_name.endswith(".xlsx"):
            raise RuntimeError("cannot save file as .xlsx,please use .xls")
        self.book_write_obj.save(file_name)


class Excel(object):

    def xls_to_xlsx(cls, file_path):

        excel_o = ReadExcel()
        excel_o.open_file(file_path)
        sheet_names = excel_o.get_sheet_names()

        wb = Workbook()
        ws = wb.active

        for sheet_index, sheet_name in enumerate(sheet_names):

            if sheet_index == 0:
                ws.title = sheet_name
            else:
                ws = wb.create_sheet(sheet_name)

            excel_o.set_active_sheet(sheet_index + 1)
            for row in excel_o.get_multiple_row_data():
                ws.append(row)

        name_list = file_path.split(".")
        name_list[len(name_list) - 1] = "xlsx"
        file_path = ".".join(name_list)

        wb.save(file_path)

    def xlsx_to_xls(cls, file_path):

        excel_o = ReadExcel()
        excel_o.open_file(file_path)
        sheet_names = excel_o.get_sheet_names()

        excel_w = WriteExcel()
        excel_w.create_file()

        for sheet_index, sheet_name in enumerate(sheet_names):

            excel_w.add_sheet(sheet_name)

            excel_o.set_active_sheet(sheet_index + 1)
            for row_index, row in enumerate(excel_o.get_multiple_row_data()):
                for column_index, data in enumerate(row):
                    excel_w.set_cell_data(row_index + 1,
                                          column_index + 1,
                                          data)

        name_list = file_path.split(".")
        name_list[len(name_list) - 1] = "xls"
        file_path = ".".join(name_list)

        excel_w.save_as(file_path)

    def motify_sheet_name(cls, file_path, sheet_index, new_sheet_name):
        from openpyxl import load_workbook
        wb = load_workbook(file_path)
        sheet_names = wb.get_sheet_names()
        ws = wb[sheet_names[sheet_index - 1]]
        ws.title = new_sheet_name
        wb.save(file_path)
        wb.close()
