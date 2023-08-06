# coding: utf-8
from __future__ import unicode_literals

import shutil
from collections import OrderedDict
from io import BytesIO
from tempfile import NamedTemporaryFile

from babel.messages.catalog import Catalog
from babel.messages.pofile import write_po, read_po
from django.utils import timezone, translation
from django.utils.encoding import smart_bytes
from modeltranslation.translator import translator
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Protection, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.protection import SheetProtection

from .import_translation import normalize_text
from .settings import (EXPORT_FILTERS, DEFAULT_TO_LANG, DEFAULT_FROM_LANG)
from .utils import get_cleaned_fields, parse_model, get_opts_from_model
from .utils.xml.render import XMLRenderer

UNTRANSLATED = 'U'
TRANSLATED = 'T'


def build_comment(instance):
    meta = instance._meta
    return '{app_title}::{model_title}:{obj} [{obj.id}]'.format(
        app_title=meta.app_config.verbose_name,
        model_title=meta.verbose_name,
        obj=instance
    )


def allow_export(msg_str, msg_id, translate_status=None):
    if not msg_id:
        return False
    if not translate_status:
        return True
    if translate_status == UNTRANSLATED and not msg_str:
        return True
    if translate_status == TRANSLATED and msg_str:
        return True

    return False


def filter_queryset(queryset, model_opts, export_filters=EXPORT_FILTERS):
    if not export_filters:
        return queryset
    for k in [model_opts['model_key'], None]:
        filter_cb = export_filters.get(k)
        if filter_cb and callable(filter_cb):
            return filter_cb(queryset, model_opts)
    return queryset


def collect_queryset_translations(qs, fields=None):
    model = qs.model
    model_opts = get_opts_from_model(model)
    trans_opts = model_opts['opts']

    fields = set(map(lambda f: f.split('.')[-1], fields or []))

    trans_fields = OrderedDict([
        [f_name, v] for f_name, v in sorted(model_opts['fields'].items(), key=lambda v: v[0])
        if not fields or f_name in fields
    ])

    for o in qs.distinct():
        for f, trans_f in trans_fields.items():
            translated_data = {lang: normalize_text(getattr(o, tf)) for lang, tf in trans_f.items()}
            context = None
            if hasattr(trans_opts, 'get_context'):
                context = trans_opts.get_context(o)

            if hasattr(trans_opts, 'get_comment'):
                comment = trans_opts.get_comment(o)
            else:
                comment = build_comment(o)

            yield dict(
                model_key=model_opts['model_key'],
                model_name=model_opts['model_name'],
                object_id=str(o.pk),
                field=f,
                model=model,
                obj=o,
                translated_data=translated_data,
                context=context,
                comment=comment
            )


def collect_model_translations(model_opts, fields=None):
    """
    :param model_opts:
    :param fields: list of field_name
    :return: iterator
    """
    model = model_opts['model']
    qs = model.objects
    if hasattr(model_opts['opts'], 'get_queryset'):
        qs = model_opts['opts'].get_queryset()
    qs = filter_queryset(qs, model_opts)
    return collect_queryset_translations(qs, fields)


def collect_models(includes=None, excludes=None):
    """
    :param models: app_label | app_label.Model | app_label.Model.field
    :param excludes: list of app_label | app_label.Model | app_label.Model.field
    :return: iterator of {model_key, model, fields, app_label, model_name}
    """
    models = translator.get_registered_models(abstract=False)
    excludes = excludes and map(parse_model, excludes)
    includes = includes and map(parse_model, includes)

    for model in models:
        only_fields = get_cleaned_fields(model, includes=includes, excludes=excludes)
        if not only_fields:
            continue
        yield get_opts_from_model(model, only_fields)


def collect_translations(
        from_lang=DEFAULT_FROM_LANG,
        to_lang=DEFAULT_TO_LANG,
        translate_status=None,
        includes=False,
        excludes=False,
        queryset=None,
):
    with translation.override(from_lang):
        if queryset:
            translations = collect_queryset_translations(queryset, fields=includes)
        else:
            collected_models = collect_models(includes, excludes)
            translations = (
                t for model_opts in collected_models
                for t in collect_model_translations(model_opts)
            )
    for tr in translations:
        msg_id = tr['translated_data'][from_lang]
        msg_str = tr['translated_data'][to_lang]
        if not allow_export(msg_str, msg_id, translate_status):
            continue
        yield tr


def export_po(translations,
              from_lang=DEFAULT_FROM_LANG,
              to_lang=DEFAULT_TO_LANG,
              queryset=None,
              stream=None,
              ):
    stream = stream or BytesIO()
    assert hasattr(stream, 'write') and hasattr(stream, 'seek'), "stream must be file-like object"
    catalog = Catalog(locale=to_lang)
    for tr in translations:
        msg_location = ('{model_key}.{field}.{object_id}'.format(**tr), 0)

        msg_id = tr['translated_data'][from_lang]
        msg_str = tr['translated_data'][to_lang]

        context = tr['context']
        comments = [tr['comment']]
        catalog.add(msg_id, msg_str, locations=(msg_location,),
                    auto_comments=comments, context=context)

    if queryset:
        # Особая уличная магия,
        # для корректной выгрузки переводов одного объекта,
        # но он может быть переведен в других объектах
        opts = get_opts_from_model(queryset.model)
        qs_locations = {"%s.%s" % (opts['model_key'], pk) for pk in
                        queryset.values_list('pk', flat=True)}
        new_catalog = Catalog(locale=to_lang)
        for message in catalog:
            locations = set()

            for (loc, n) in message.locations:
                spl = loc.split('.')
                del spl[2]
                locations.add(".".join(spl))

            if locations & qs_locations:
                kw = {k: getattr(message, k) for k in ['auto_comments', 'locations', 'context']}

                new_catalog.add(
                    message.id,
                    message.string,
                    **kw
                )
        catalog = new_catalog

    write_po(stream, catalog)
    stream.seek(0)
    return stream


def export_xlsx(translations,
                from_lang=DEFAULT_FROM_LANG,
                to_lang=DEFAULT_TO_LANG,
                queryset=None,
                stream=None,
                ):
    stream = stream or BytesIO()
    po_file_stream = export_po(
        translations=translations, from_lang=from_lang,
        to_lang=to_lang, queryset=queryset)

    catalog = read_po(po_file_stream)

    wb = Workbook()
    ws = wb.active
    ws.protection = SheetProtection(sheet=True,
                                    selectLockedCells=False,
                                    selectUnlockedCells=False,
                                    formatColumns=False,
                                    formatRows=False,
                                    )
    ws.append(["comment", "locations", from_lang, to_lang])
    align = Alignment(wrap_text=True, vertical='top')
    for i, m in enumerate(catalog, start=1):
        if not m.id:
            continue
        comment = ""
        if m.context:
            comment = f"Context: {m.context}\n"
        comment += "\n".join(m.auto_comments)
        comment_cell = WriteOnlyCell(ws, value=comment)
        from_lang_cell = WriteOnlyCell(ws, value=m.id)
        to_lang_cell = WriteOnlyCell(ws, value=normalize_text(m.string))
        to_lang_cell.protection = Protection(locked=False)

        comment_cell.alignment = align
        from_lang_cell.alignment = align
        to_lang_cell.alignment = align

        locations = '\n'.join([path for path, _ in m.locations])
        meta_cell = WriteOnlyCell(ws, value=locations)
        max_height_lines = max(
            [len(c.value.splitlines()) for c in [comment_cell, from_lang_cell, to_lang_cell] if
             c.value])
        ws.append([comment_cell, meta_cell, from_lang_cell, to_lang_cell])
        ws.row_dimensions[i].height = 20 * max_height_lines

    ws.column_dimensions['B'].hidden = True

    ws.column_dimensions[get_column_letter(1)].width = 50
    for c in range(2, 5):
        ws.column_dimensions[get_column_letter(c)].width = 100

    with NamedTemporaryFile(suffix=".xlsx") as tf:
        wb.save(tf.name)

        shutil.copyfileobj(tf.file, stream)

    stream.flush()
    stream.seek(0)
    return stream


def export_xml(translations,
               from_lang=DEFAULT_FROM_LANG,
               to_lang=DEFAULT_TO_LANG,
               queryset=None,
               stream=None,
               merge_trans=True,
               ):
    stream = stream or BytesIO()

    xml_translations = []
    if merge_trans:
        po_file_stream = export_po(
            translations=translations, from_lang=from_lang,
            to_lang=to_lang, queryset=queryset)
        catalog = read_po(po_file_stream)
        for c in catalog:
            if not c.id:
                continue

            t = {
                '@id': ';'.join([c[0] for c in c.locations]),
                '@comment': ';'.join(c.auto_comments),
                'lang': [
                    {'@code': from_lang, '#text': c.id},
                    {'@code': to_lang, '#text': c.string},
                ]
            }
            if c.context:
                t['@context'] = c.context
            xml_translations.append(t)
    else:
        objects = {}
        for tr in translations:
            if tr['object_id'] not in objects:
                objects[tr['object_id']] = {
                    '@id': '{model_key}.{object_id}'.format(**tr),
                    '@comment': tr['comment'],
                    'fields': [],
                }
                if tr['context']:
                    objects[tr['object_id']]['@context'] = tr['context']

            objects[tr['object_id']]['fields'].append({
                tr['field']: {
                    '@field': tr['field'],
                    'lang': [
                        {'@code': from_lang, '#text': tr['translated_data'][from_lang]},
                        {'@code': to_lang, '#text': tr['translated_data'][to_lang] or ''},
                    ]
                }
            })
        xml_translations = list(objects.values())

    xml = XMLRenderer()
    root_attrs = {
        'created': timezone.now().isoformat()
    }
    xml_t = xml.render(xml_translations,
                       parent_tag_name='object',
                       root_attrs=root_attrs)
    stream.write(smart_bytes(xml_t))

    stream.flush()
    stream.seek(0)
    return stream
