import pandas as pd
from functools import partial
from difflib import SequenceMatcher
import hashlib
import os
from openpyxl.utils import get_column_letter

import logging
import click

from halo import Halo
from hycli.commons import calculate_improvements as improvements
import datetime


class ModelComparer:
    """A class used for performing models comparison.

    Contains several methods of loading files, checking files, comparing differences and exporting comparison to Excel.

    Attributes:
            file_path_1 (str): filepath to the ground truth in excel format
            file_path_2 (str): filepath to the model extraction in excel format

    """

    def __init__(self, file_path_1: str, file_path_2: str):
        self.file_path_1 = file_path_1
        self.file_path_2 = file_path_2

    def load_files(self):
        """Performs loading objects into the pandas dataframes from excel structure."""

        def _load_file(file_path: str) -> dict:
            def _check_extesion() -> str:
                file_name = os.path.basename(file_path)
                extension = os.path.splitext(file_name)[1]
                return extension

            if _check_extesion()[:3] == ".xl":
                return dict(pd.read_excel(io=file_path, sheet_name=None, engine="openpyxl"))

            else:
                raise TypeError(
                    f"{file_path}: Type {_check_extesion()} is currently not supported.\nSupported types: xlsx, xls, "
                    f"xlsm"
                )

        self.model_1 = _load_file(self.file_path_1)
        self.model_2 = _load_file(self.file_path_2)

    def check_files(self):
        """Checks dataframes"""

        def _check_dataframe(df, ws_name, path):
            """Checks if dataframe has a column named file_name which is required to perform other calculations"""
            # noqa
            assert "file_name" in df.columns, (
                f"Worksheet '{ws_name}' in file '{path}' has no column 'file_name' "
                f"which is required to perform further operations. \n "
                f"Please check the names of columns and run the program once again."
            )

        [
            _check_dataframe(df, name, self.file_path_1)
            for name, df in self.model_1.items()
        ]
        [
            _check_dataframe(df, name, self.file_path_2)
            for name, df in self.model_2.items()
        ]

    @staticmethod
    def export_comparison_to_excel(comparisons: dict, file_path: str):
        """Exports comparison to excel file.

        Args:
            comparisons (dict): dict containing comparisons stored in pandas dataframes
            file_path (str): output file path of comparison in Excel format

        """

        def _format_excel(df: pd.DataFrame) -> pd.DataFrame:
            """Takes pandas dataframe and formats excel structure

            Saves dataframes into specific worksheets;
            Sets colours indicating differences;
            Sets column widths.

            Args:
                df: pandas dataframe

            Returns:
                styled pandas dataframe

            """

            def _highlight_records(val):
                val = -1 if val == "" else val

                if val > 0:
                    color = "#FFB6C1"
                    return f"background-color: {color}"

            diff_cols = [col for col in df.columns if "diff_" in col]

            df = (
                df.reset_index(drop=True)
                .style.set_properties(**{"background-color": "#FFFFFF"})
                .applymap(_highlight_records, subset=list(diff_cols))
            )

            return df

        def _set_column_widths(sheet: str):
            """Sets custom column widths for each columns in worksheet"""
            writer.sheets[sheet].column_dimensions["A"].width = 5
            writer.sheets[sheet].column_dimensions["B"].width = 25
            for x in range(3, writer.sheets[sheet].max_column + 1):
                writer.sheets[sheet].column_dimensions[get_column_letter(x)].width = 15

        with pd.ExcelWriter(file_path, engine="openpyxl", mode="w") as writer:
            for k, v in comparisons.items():
                v = _format_excel(v)
                v.to_excel(writer, sheet_name=k)
                _set_column_widths(k)
            writer.save()

    # TODO: this function should be simplified
    def compare_differences(self) -> dict:  # noqa
        def _apply_difference(s: pd.DataFrame, a: str, b: str, tolerance=0.05) -> float:
            """
            Applies difference ratio (from 0 to 1) between each record on given pandas DataFrame
            using Levenshtein distance.
            Difference is calculated between x column and y column.
            """

            try:
                # if both numbers are equal then return difference = 0
                if s[a] == s[b]:
                    difference = 0
                    return difference

                # if both datatypes are null return difference = 0
                if pd.isna(s[a]) and pd.isna(s[b]):
                    difference = 0
                    return difference

                else:
                    # if one of the datatypes is null then treat is as a 0
                    if pd.isna(s[a]):
                        s[a] = 0

                    if pd.isna(s[b]):
                        s[b] = 0

                    # if one of the datatypes is datetime then treat is as a strformat
                    if isinstance(s[a], datetime.datetime):
                        s[a] = s[a].strftime("%Y-%m-%d")

                    if isinstance(s[b], datetime.datetime):
                        s[b] = s[b].strftime("%Y-%m-%d")

                    # For comparing numbers
                    if all(
                        [
                            isinstance(s[a], (int, float, bool)),
                            isinstance(s[b], (int, float, bool)),
                        ]
                    ):
                        higher_value = float(max([s[a], s[b]]))
                        lower_value = float(min([s[a], s[b]]))
                        if lower_value == 0:
                            lower_value = 0.01

                        difference = round(abs(1 - higher_value / lower_value), 2)

                        if difference > tolerance:
                            difference = 1

                    # For comparing strings
                    else:

                        similarity = SequenceMatcher(
                            None, str(s[a]).lower(), str(s[b]).lower()
                        ).ratio()

                        if similarity < 0.6:
                            difference = 1
                        else:
                            difference = round(1.00 - similarity, 2)

                    return difference
            except KeyError:
                # In case there is no counterpart column
                return 1

        def _create_difference_for_each_column(df: pd.DataFrame) -> pd.DataFrame:
            for _ in df.columns:

                if _[-2:] == "_x":  # for each _x column (ground truth)
                    x_column = _  # column which has _x suffix (from ground truth)
                    new_column = (
                        "diff_" + _[:-2]
                    )  # column which has diff_ prefix with calculated difference
                    y_column = (
                        _[:-2] + "_y"
                    )  # column which has _y suffix (some dataset)

                    # Calculate Levenshtein difference between each record in column x and column y
                    df[new_column] = df.apply(
                        partial(_apply_difference, a=x_column, b=y_column), axis=1
                    )

            return df

        def _diff_summary_by_row(df: pd.DataFrame) -> pd.DataFrame:
            """Provide difference summary for each row"""
            arr = []
            for column in df.columns:
                if column[:5] == "diff_":
                    arr.append(column)
            return df[arr].sum(axis=1)

        def _sort_columns(df: pd.DataFrame) -> list:
            """Use this function to sort columns with order x, y, difference"""
            arr = ["file_name", "distance", "diff_total"]
            for _ in df.columns:

                if _[-2:] == "_x":
                    arr.append(_)
                    arr.append(_[:-2] + "_y")
                    arr.append("diff_" + _[:-2])

            return arr

        def _summary(df_1: pd.DataFrame, df_2: pd.DataFrame) -> pd.DataFrame:
            """
            Provide summary based on original df and results.
            df_1: ground_truth
            df_2: extraction
            """
            number_of_documents = df_1.shape[0]
            item_categories = len(
                [
                    column
                    for column in df_2.columns
                    if column.startswith("diff_") and column != "diff_total"
                ]
            )
            data_points = number_of_documents * item_categories
            percent_diff = round(self.total_difference / data_points, 2)

            data = {
                1: ["Number of documents", None, number_of_documents],
                2: ["Number of compared columns", None, item_categories],
                3: ["Compared data points", None, data_points],
                4: ["Overall difference in percent", None, percent_diff],
                5: ["Overall accuracy in percent", None, 1 - percent_diff],
            }

            summary = pd.DataFrame.from_dict(
                data, orient="index", columns=["file_name", "distance", "diff_total"]
            )

            df_2 = df_2.append(summary)

            return df_2

        def _compare_worksheets(
            left_dataframe_name: str, right_dataframe_name: str
        ) -> pd.DataFrame:
            def _apply_best_matches(
                df_left: pd.DataFrame,
                df_right: pd.DataFrame,
                file_name: str,
                left_matching_index="file_name",
                right_matching_index="file_name",
            ) -> pd.DataFrame:
                """
                Based on two dataframes, perform left-join type matching using Levenshtein lookup for each row.
                """

                filtered_file_name = df_left[df_left.file_name == file_name].copy()

                # prepare indexes for artificial_keys
                filtered_file_name["index_x"] = [
                    hashlib.md5(val.encode("utf-8")).hexdigest()
                    for val in filtered_file_name["artificial_key"]
                ]
                df_right["index_y"] = [
                    hashlib.md5(val.encode("utf-8")).hexdigest()
                    for val in df_right["artificial_key"]
                ]

                # create cartesian product for file
                cartesian = pd.merge(
                    filtered_file_name,
                    df_right,
                    left_on=left_matching_index,
                    right_on=right_matching_index,
                )

                if not cartesian.empty:

                    # how many records we have on _x
                    records_on_x = filtered_file_name["file_name"].count()
                    # how many records we have on _y
                    records_on_y = df_right[df_right.file_name == file_name][
                        "file_name"
                    ].count()

                    cartesian["distance"] = cartesian.apply(
                        partial(
                            _apply_difference,
                            a="artificial_key_x",
                            b="artificial_key_y",
                        ),
                        axis=1,
                    )

                    def _sudoku_matcher(
                        df: pd.DataFrame, per_index: str
                    ) -> pd.DataFrame:
                        """

                        Creates a dataframe based on a list of matches that are possible (e.g. matching A1 to A5 and
                        A2 to A4 is possible, while matching A1 to A5 and A1 to A4 is not possible due to A1 being
                        matched twice)

                        """

                        # introduce empty DataFrame at the beginning
                        best_results = pd.DataFrame()

                        # copy a dataset that we will working with
                        dataset = df

                        # execute in loop until dataset would have no possible matches
                        while not dataset.empty:
                            # collect the best match
                            best_match = dataset.loc[dataset["distance"].idxmin()]
                            best_results = best_results.append(best_match)

                            index_to_be_removed = best_match.index_x
                            corresponding_field = best_match.index_y

                            # perform elimination within dataset
                            dataset = dataset.drop(
                                dataset[dataset["index_x"] == index_to_be_removed].index
                            )
                            dataset = dataset.drop(
                                dataset[dataset["index_y"] == corresponding_field].index
                            )

                        # TODO: this should be moved to unit tests
                        # assertions checks that there should be no duplicates after the executing the loop
                        # within matched dataset

                        assert not dataset["index_y"].duplicated().any(), dataset
                        assert not dataset["index_x"].duplicated().any(), dataset
                        assert not best_results["index_x"].duplicated().any(), (
                            file_name,
                            best_results.index_x,
                        )
                        assert not best_results["index_y"].duplicated().any(), (
                            file_name,
                            best_results.index_y,
                        )

                        # if there are more items on a left dataset than in the right dataset
                        # collect not matched records from left dataset (x) and place them in a results as non matched
                        if per_index == "index_x":
                            not_matched_results = df[
                                ~df.index_x.isin(best_results.index_x)
                            ]
                            if not not_matched_results.empty:
                                dummy_indexes = not_matched_results["index_x"].unique()
                                dummy_records = filtered_file_name[
                                    filtered_file_name.index_x.isin(dummy_indexes)
                                ]
                                old_columns = dummy_records.columns
                                new_columns = [
                                    column + "_x"
                                    if (column != "file_name" and column != "index_x")
                                    else column
                                    for column in old_columns
                                ]
                                dummy_records.columns = new_columns
                                best_results = best_results.append(dummy_records)

                        return best_results

                    matches = (
                        _sudoku_matcher(cartesian, "index_x")
                        if records_on_x >= records_on_y
                        else _sudoku_matcher(cartesian, "index_y")
                    )
                    matches.drop(["artificial_key_x"], axis=1, inplace=True)
                    matches.drop(["artificial_key_y"], axis=1, inplace=True)
                    matches.drop(["index_x"], axis=1, inplace=True)
                    matches.drop(["index_y"], axis=1, inplace=True)

                    return matches

                else:
                    matches = pd.merge(
                        filtered_file_name, df_right, on="file_name", how="left"
                    )
                    matches.drop(["artificial_key_x"], axis=1, inplace=True)
                    return matches

            # Prepare data
            df_1 = self.model_1[left_dataframe_name]
            df_2 = self.model_2[right_dataframe_name]

            def _perform_left_join(
                df_left: pd.DataFrame, df_right: pd.DataFrame
            ) -> pd.DataFrame:
                """Performs left join based on two dataframes

                Args:
                    df_left: left dataset
                    df_right: right dataframe

                Returns:
                    pandas dataframe

                """
                df = pd.merge(
                    df_left, df_right, left_on="file_name", right_on="file_name"
                )
                return df

            if any(df_1["file_name"].duplicated().to_list()):
                comparison_type = "multi"
            else:
                comparison_type = "single"

            df_1["artificial_key"] = df_1.astype(str).values.sum(axis=1)
            df_2["artificial_key"] = df_2.astype(str).values.sum(axis=1)
            df_1["file_name"] = df_1["file_name"].astype(str)
            df_2["file_name"] = df_2["file_name"].astype(str)

            grouped_1 = df_1.groupby("file_name")["file_name"].count()

            if comparison_type == "single":
                results = _perform_left_join(df_left=df_1, df_right=df_2)
            else:
                # Apply best matches for each file
                results = pd.DataFrame()
                for k, v in grouped_1.to_dict().items():
                    r = _apply_best_matches(df_1, df_2, k)
                    results = results.append(r)

            results = _create_difference_for_each_column(results)
            return results

        def _line_items_statistics(dataframes: tuple) -> pd.DataFrame:

            df_1 = dataframes[0]
            df_2 = dataframes[1]

            group_1 = df_1.groupby(["file_name"])["row_number"].count()
            group_2 = df_2.groupby(["file_name"])["row_number"].count()

            df = pd.merge(
                group_1, group_2, how="outer", left_on="file_name", right_on="file_name"
            )
            df.fillna(0, inplace=True)
            df["diff_count_of_items"] = df["row_number_x"] - df["row_number_y"]
            df.rename(
                columns={
                    "row_number_x": "count_of_x_items",
                    "row_number_y": "count_of_y_items",
                },
                inplace=True,
            )
            df["count_of_x_items"].astype("int")
            df["count_of_y_items"].astype("int")

            return df

        def run_excel_flow() -> dict:
            comparisons = {}

            def _match_worksheets_by_column_names() -> list:
                """Matches the worksheets looking for closest matches using column names in both workbooks,
                disregard worksheet names.

                Returns:
                    list: a list of tuples representing worksheet matching
                """
                columns_in_workbook_1 = {
                    worksheet_name: df.columns.to_list()
                    for worksheet_name, df in self.model_1.items()
                }
                columns_in_workbook_2 = {
                    worksheet_name: df.columns.to_list()
                    for worksheet_name, df in self.model_2.items()
                }

                def _get_best_match(
                    columns_in_df_1: list, propositions: dict, threshold: int = 2
                ):
                    """Matches the best combination of column names within a second workbook.

                    Args:
                        columns_in_df_1: represents a list of columns in first workbook
                        propositions: represents a key value pairs of worksheet names and columns of a second workbook
                        threshold: represents how many records should be included in second worksheet to consider it
                        as a match.

                    Returns: string or None

                    """
                    scores = {}
                    for ws_name, columns in propositions.items():
                        scores[ws_name] = set(columns_in_df_1).intersection(
                            set(columns)
                        )

                    if len(max(scores.values())) < threshold:
                        return None

                    best_match = max(scores, key=scores.get)

                    return best_match

                pair_of_matched_worksheets = []
                for ws_name, columns_in_workbook_1 in columns_in_workbook_1.items():
                    if columns_in_workbook_2:
                        closest_match = _get_best_match(
                            columns_in_workbook_1, columns_in_workbook_2
                        )

                        if closest_match:
                            pair_of_matched_worksheets.append((ws_name, closest_match))
                            columns_in_workbook_2.pop(closest_match)
                    else:
                        break

                return pair_of_matched_worksheets

            matched_worksheets = _match_worksheets_by_column_names()

            if not matched_worksheets:
                raise Exception(
                    "The algorithm has no possibility to match any worksheet from second workbook. "
                    "Please check the columns names in both files and try again."
                )

            for pair_of_worksheets in matched_worksheets:
                comparison_results = _compare_worksheets(
                    pair_of_worksheets[0], pair_of_worksheets[1]
                )

                # Prepare total for each row
                comparison_results["diff_total"] = _diff_summary_by_row(
                    comparison_results
                )

                # Sort columns
                comparison_results = pd.DataFrame(
                    comparison_results, columns=_sort_columns(comparison_results)
                )

                diff_cols = [
                    col for col in comparison_results.columns if ("diff_" in col)
                ]

                # Calculate total difference in percent
                number_of_records = self.model_1[pair_of_worksheets[0]].shape[0]

                # Calculate total difference
                sum_of_differences = round(
                    comparison_results[diff_cols].sum(numeric_only=True, axis=0)
                )
                percent_differences = round(sum_of_differences / number_of_records, 2)

                # Place total difference
                comparison_results.loc["diff_total"] = round(sum_of_differences, 2)

                # Place below percent difference
                comparison_results.loc["diff_percent"] = percent_differences

                comparison_results.at["diff_total", "file_name"] = "Total difference"
                comparison_results.at[
                    "diff_percent", "file_name"
                ] = "Total difference in percent"

                self.total_difference = comparison_results.loc[
                    "diff_total", "diff_total"
                ]
                comparison_name = pair_of_worksheets[0]
                comparison_results = _summary(
                    self.model_1[comparison_name], comparison_results
                )

                comparisons[comparison_name] = comparison_results

            try:
                comparisons["line_items_statistics"] = _line_items_statistics(
                    (self.model_1["items"], self.model_2["items"])
                )
            except Exception:
                logging.warning(
                    "Skipping preparing line item statistics. "
                    "Please check if worksheet 'items' exists in both workbooks"
                )

            return comparisons

        return run_excel_flow()


@click.command(context_settings=dict(max_content_width=200))
@click.argument("file_path_1", type=click.Path(exists=True))
@click.argument("file_path_2", type=click.Path(exists=True))
@click.option(
    "-o",
    "--output",
    help="output path for xlsx file relative from current location (ends with .xlsx)",
)
@click.option(
    "-t",
    "--truth",
    help="ground truth path. For performing three way comparison",
    default=None,
)
@click.pass_context
def compare(ctx, file_path_1, file_path_2, output=None, truth=None):
    """ compares extraction data between two excel files """
    spinner = Halo(spinner="dots")
    spinner.start()

    def _evaluate_models():
        """Performs 3- way comparison. Evaluates which model performs better."""
        spinner.start("Starting evaluation..\n")

        mc_1 = ModelComparer(truth, file_path_1)
        mc_2 = ModelComparer(truth, file_path_2)

        mc_1.load_files()
        mc_2.load_files()

        comparisons_1 = mc_1.compare_differences()
        comparisons_2 = mc_2.compare_differences()

        evaluation_of_improvements = improvements.calculate_improvements(
            comparison_1=comparisons_1, comparison_2=comparisons_2
        )
        evaluation_of_improvements.to_excel(output, sheet_name="header_evaluation")

        spinner.succeed(f"Please check evaluation here: {output} \n")

    def _compare_models():
        """ Performs comparison between two datasets"""

        spinner.start("Starting comparison..\n")
        mc = ModelComparer(file_path_1, file_path_2)
        mc.load_files()
        comparisons = mc.compare_differences()
        mc.export_comparison_to_excel(comparisons, output)
        spinner.succeed(f"Please check comparison here: {output} \n")

    if truth:
        output = "evaluation.xlsx" if output is None else output
        _evaluate_models()
    else:
        output = "comparison.xlsx" if output is None else output
        _compare_models()
