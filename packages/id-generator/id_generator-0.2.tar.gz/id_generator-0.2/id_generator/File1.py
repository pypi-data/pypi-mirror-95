import openpyxl as op
import random
import string


class id_generate:
    def __init__(self, file_path, seed, number_of_characters, number_of_times):
        self.file_path = file_path
        self.seed = seed
        self.number_of_characters = number_of_characters
        self.number_of_times = number_of_times
        wb = op.load_workbook(r"{}".format(self.file_path))
        ws = wb.active
        ws["A1"] = "Unique ID"

        random.seed(int(self.seed))


        for i in range(self.number_of_times):
            ws[f"A{i+2}"] =  "".join(random.choice(string.ascii_uppercase + string.digits) for _ in range(self.number_of_characters))
        
        wb.save(r"{}".format(self.file_path))

        return

    

