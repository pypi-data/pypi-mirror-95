#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Author : mocobk
# @Email : mailmzb@qq.com
# @Time : 2021/2/19 18:20
import pickle
from collections import OrderedDict
from pathlib import Path
from urllib.parse import urljoin
from zipfile import ZipFile, ZIP_DEFLATED

import requests
from openpyxl import load_workbook
from openpyxl.utils import quote_sheetname
from openpyxl.worksheet.datavalidation import DataValidation

ROOT_DIR = Path(__file__).parent
TEMPLATE_XLS = ROOT_DIR.joinpath('template', 'case.xlsm')
TAPD_BASE_URL = 'http://9.134.187.9:8080'

BVT_TASK_TEMPLATE = ROOT_DIR.joinpath('template', 'bvt_task.html')


def copy_drawings_from_xlsx(source, target):
    """
    已弃用
    openpyxl 处理的 Excel 会丢失掉形状等元素
    https://openpyxl.readthedocs.io/en/latest/usage.html?highlight=keep_vba#write-a-workbook-from-xltm-as-xlsm
    为了保持模板的 形状按钮 不丢失，这里采用了很差劲的方法，把源文件的 drawings 移到新生产的 xlsx
    目前写死了只还原 sheet1 的内容
    注意：源文件至少要存在一个控件，以便生成相应的关联，未找到原因
    xl/worksheets/sheet1.xml.rels drawing1 关联文件
    xl/drawings/drawing1.xml 绘图文件
    """
    source_archive = ZipFile(source, 'r')
    target_archive = ZipFile(target, 'r')
    sheet1 = 'xl/worksheets/sheet1.xml'
    copy_files = ['xl/worksheets/_rels/sheet1.xml.rels', 'xl/drawings/drawing1.xml']
    tmp = {}

    for file in target_archive.namelist():
        if file not in copy_files:
            tmp[file] = target_archive.read(file)

    tmp[sheet1] = target_archive.read(sheet1).replace(b'<legacyDrawing r:id="anysvml" /></worksheet>',
                                                      b'<drawing r:id="rId1"/><legacyDrawing r:id="rId2"/></worksheet>')
    target_archive.close()

    for file in source_archive.namelist():
        if file in copy_files:
            tmp[file] = source_archive.read(file)
    source_archive.close()

    output_archive = ZipFile(target, 'w', ZIP_DEFLATED, allowZip64=True)

    for file in tmp:
        output_archive.writestr(file, tmp[file])

    output_archive.close()


def secure_filename(filename: str):
    chars_map = {'\\': '_',
                 '/': '_',
                 ':': '_',
                 '*': '_',
                 '?': '',
                 '"': '',
                 '<': '_',
                 '>': '_',
                 '|': '_'}

    trans_table = filename.maketrans(chars_map)
    return filename.translate(trans_table)


class Config:
    CONFIG_FILE = ROOT_DIR.joinpath('config.pkl')

    def get(self):
        if self.CONFIG_FILE.exists():
            return pickle.load(self.CONFIG_FILE.open('rb'))
        else:
            return {}

    def put(self, config):
        pickle.dump(config, self.CONFIG_FILE.open('wb'))


class Tapd:
    def __init__(self, iteration_id=None, base_url=None):
        self.base_url = base_url or TAPD_BASE_URL
        self.iteration_id = iteration_id

    def set_iteration_id(self, iteration_id):
        self.iteration_id = iteration_id

    def get_tapd_iterations(self):
        url = urljoin(self.base_url, '/tapd/origin/iterations')
        response = requests.get(url, params={'status': 'open', 'fields': 'id,name'})
        data = [(item['Iteration']['name'], item['Iteration']['id']) for item in response.json().get('data', [])]
        return OrderedDict(data[:8])

    def get_tapd_iteration_developers(self):
        url = urljoin(self.base_url, '/tapd/origin/tasks')
        response = requests.get(
            url, params={'iteration_id': self.iteration_id, 'fields': 'owner', 'order': 'owner asc'}
        )
        data = [item['Task']['owner'] for item in response.json().get('data', []) if item['Task']['owner']]
        return set(data)

    def get_tapd_iteration_stories(self):
        url = urljoin(self.base_url, '/tapd/origin/stories')
        response = requests.get(url, params={'iteration_id': self.iteration_id, 'fields': 'id,name'})
        data = [(item['Story']['name'], item['Story']['id']) for item in response.json().get('data', [])]
        return OrderedDict(data)

    def get_tapd_case_categories(self):
        url = urljoin(self.base_url, '/tapd/origin/tcase_categories')
        response = requests.get(url, params={'parent_id': 0, 'fields': 'id,name', 'order': 'name asc'})
        data = [(item['TcaseCategory']['name'], item['TcaseCategory']['id']) for item in
                response.json().get('data', [])]
        return OrderedDict(data)

    def create_tapd_task(self, **kwargs):
        """name, description, creator, owner, story_id, iteration_id"""
        url = urljoin(self.base_url, '/tapd/origin/tasks')
        kwargs.setdefault('iteration_id', self.iteration_id)
        kwargs.setdefault('priority', 4)
        kwargs.setdefault('custom_field_one', '冒烟')
        response = requests.post(url, data=kwargs)
        task_id = response.json().get('data', {}).get('Task', {}).get('id')
        return task_id


class Excel:
    def __init__(self, filename, tapd: Tapd = None):
        self.filename = filename
        self.excel_wb = load_workbook(filename, keep_vba=True)
        self.excel_case_ws = self.excel_wb.active
        self.excel_validate_ws = self.excel_wb.create_sheet('数据验证（勿删）')
        self.title_map = self.get_title_map()
        self.tapd = tapd

    def get_title_map(self):
        title_map = {}
        for (cell,) in self.excel_case_ws.iter_cols(min_col=1, max_col=50, min_row=2, max_row=2):
            if cell.value is None:
                break
            title_map[cell.value] = cell
        return title_map

    def set_sheet_name(self, name):
        self.excel_case_ws.title = name

    @staticmethod
    def append_cols(sheet, column_data):
        column_index = cur_max_column = sheet.max_column
        if sheet.cell(1, cur_max_column).value is not None:
            column_index += 1
        for index, value in enumerate(column_data, start=1):
            sheet.cell(index, column_index, value)

    def set_data_validation(self, cell, formula_data, show_error_message=True):
        self.append_cols(self.excel_validate_ws, formula_data)
        column_letter = self.excel_validate_ws.cell(1, self.excel_validate_ws.max_column).column_letter
        dv = DataValidation(
            type="list",
            formula1=f"{quote_sheetname(self.excel_validate_ws.title)}!${column_letter}$2:${column_letter}${len(formula_data)}",
            allow_blank=True,
            showErrorMessage=show_error_message)
        dv.error = '请使用下拉选择可用的值'
        dv.errorTitle = '输入的值有误'
        dv.add(cell)
        self.excel_case_ws.add_data_validation(dv)

    def set_categories_data_validation(self):
        """设置 一级模块（用例目录）的数据验证"""
        column_letter = self.title_map['一级模块'].column_letter
        cell = f'{column_letter}3:{column_letter}1048576'
        self.set_data_validation(cell, ('一级模块', *self.tapd.get_tapd_case_categories().keys()))

    def set_developer_data_validation(self):
        """设置 开发人员 的数据验证"""
        column_letter = self.title_map['开发人员'].column_letter
        cell = f'{column_letter}3:{column_letter}1048576'
        self.set_data_validation(cell, ('开发人员', *self.tapd.get_tapd_iteration_developers()), show_error_message=False)

    def set_stories_data_validation(self):
        """设置 需求 的数据验证"""
        column_letter = self.title_map['需求'].column_letter
        cell = f'{column_letter}3:{column_letter}4'
        self.set_data_validation(cell, ('需求', *self.tapd.get_tapd_iteration_stories().keys()))
        self.append_cols(self.excel_validate_ws, ('需求ID', *self.tapd.get_tapd_iteration_stories().values()))

    def save(self, filename):
        self.excel_wb.save(filename)


if __name__ == '__main__':
    pass
    # tapd = Tapd(TAPD_BASE_URL)
    # tapd.set_iteration_id('1020373562001261837')
    # # print(tapd.get_tapd_iterations())
    # # print(tapd.get_tapd_iteration_developers())
    # # print(tapd.get_tapd_iteration_stories())
    # print(tapd.get_tapd_case_categories())
    # template = Template(BVT_TASK_TEMPLATE.read_text())
    # desc = template.render(precondition='111\n222', case_steps='333\n\n444', expected='555')
    # print(tapd.create_tapd_task(name='测试任务', description=desc, creator='aidenmo;', owner='aidenmo;', story_id='1020373562862462103'))
