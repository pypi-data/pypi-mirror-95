"""Script that build and save a database postal code to circonscription.

(Circonscription is a french word for district).

Needed data:
    - districts (french: circonscription), downloaded from
    https://www.data.gouv.fr/fr/datasets/circonscriptions-legislatives-table-de-correspo
    ndance-des-communes-et-des-cantons-pour-les-elections-legislatives-de-2012-et-sa-mis
    e-a-jour-pour-les-elections-legislatives-2017/
    and saved to insee_code_to_district.xlsx
    - Insee code / postal code, downloaded from
    https://www.data.gouv.fr/fr/datasets/base-officielle-des-codes-postaux/#_
    and saved to insee_postal.csv

External sources:
    - http://www.elections-legislatives.fr/circonscriptions/075.asp?#lis
    which provided us the mean to know precisely wich arrondissement matched
    which circonscription for Paris and Marseille, and
    https://medium.com/@filgb/la-carte-des-circonscriptions-de-lyon-et-du-rh%C3%B4ne-pour-les-l%C3%A9gislatives-9a29d969c8e7
    for Lyon.

Output:
    - postal code to district number dict, saved as json object into
      postal_code_to_district.json

Dependencies:
    - openpyxl
"""
import csv
import json
import logging
from collections import defaultdict
from typing import Iterable, Tuple, Dict, DefaultDict, List, Set

from openpyxl import load_workbook

RawDistricts = Tuple
CodedDistrict = str

logger = logging.getLogger(__name__)

# Raw data : we store here data from cities where the districts better correspond to
# postal codes than to insee codes. These dict map arrondissements to their district
# numbers. The function add_static_part build the corresponding postal
# code -> district number

paris_codes = {
    1: (1,),
    2: (1,),
    3: (5,),
    4: (7,),
    5: (2,),
    6: (2, 11),
    7: (2, 12),
    8: (1,),
    9: (1,),
    10: (5,),
    11: (6, 7),
    12: (7, 8),
    13: (9, 10),
    14: (10, 11),
    15: (12, 13),
    16: (4, 14),
    116: (4, 14),
    17: (3, 4),
    18: (3, 17, 18),
    19: (16, 17, 18),
    20: (6, 8, 15),
}

marseilles_codes = {
    1: (4,),
    2: (4,),
    3: (4,),
    4: (5,),
    5: (4, 5),
    6: (4, 5),
    7: (2,),
    8: (2,),
    9: (6,),
    10: (1, 6),
    11: (1,),
    12: (1, 3),
    13: (3,),
    14: (3, 7),
    15: (7,),
    16: (7,),
}

lyon_codes = {
    1: (2,),
    2: (1, 2),
    3: (3, 4),
    4: (2,),
    5: (1,),
    6: (4,),
    7: (1, 3),
    8: (1, 3, 4),
    9: (1, 2),
}


def add_static_part(dico: DefaultDict[str, Set[CodedDistrict]]):
    """Add the static postal codes to the dico."""
    for arr, circos in paris_codes.items():
        cp = f"75{arr:03}"
        for c in circos:
            dico[cp].add(f"75.{c}")

    for arr, circos in marseilles_codes.items():
        cp = f"13{arr:03}"
        for c in circos:
            dico[cp].add(f"13.{c}")

    for arr, circos in lyon_codes.items():
        cp = f"69{arr:03}"
        for c in circos:
            dico[cp].add(f"69.{c}")


class District:
    """Represent a district (in french circonscription)."""

    code_dpt: int
    nom_dpt: str
    code_commune: int
    nom_commune: str
    code_circ_legislative: int
    code_canton: str
    nom_canton: str
    code_insee: str

    def __init__(
        self,
        code_dpt,
        nom_dpt,
        code_commune,
        nom_commune,
        code_circ_legislative,
        code_canton,
        nom_canton,
    ):
        self.code_dpt = dpt_to_int(code_dpt)
        self.nom_dpt = nom_dpt
        self.code_commune = int(code_commune)
        self.nom_commune = nom_commune
        self.code_circ_legislative = int(code_circ_legislative)
        self.code_canton = code_canton
        self.nom_canton = nom_canton

        self.key = f"{self.code_dpt}.{self.code_circ_legislative}"

        if type(code_dpt) == str and code_dpt[0] == 'Z':
            self.code_dpt, self.code_commune = dpt_commune_to_int(
                code_dpt, code_commune
            )
            self.key = "{}.{}".format(
                self.code_dpt * 10 + (self.code_commune // 100),
                self.code_circ_legislative,
            )
            self.code_insee = f"{self.code_dpt}{self.code_commune}"

        elif type(code_dpt) == str and code_dpt[1] in ('A', 'B'):
            self.code_insee = f"{code_dpt}{self.code_commune:03}"

        else:
            self.code_insee = f"{self.code_dpt:02}{self.code_commune:03}"

    def __repr__(self):
        return f"{self.nom_commune} ({self.code_dpt})"

    def __hash__(self):
        return self.key.__hash__()


def dpt_commune_to_int(dpt: str, code_commune: str) -> Tuple[int, int]:
    """For outre mer, build departement, code commune"""
    c = int(code_commune)

    if dpt == 'ZA':
        return 97, c
    if dpt == 'ZB':
        return 97, c
    if dpt == 'ZC':
        return 97, c
    if dpt == 'ZD':
        return 97, c
    if dpt == 'ZM':
        return 97, c + 100  # Because.
    if dpt == 'ZN':
        return 98, c
    if dpt == 'ZP':
        return 98, c + 700  # Again, because.
    if dpt == 'ZS':
        return 97, 502  # Only one possibility
    if dpt == 'ZW':
        return 98, 613  # Only one possibility
    if dpt == 'ZX':
        return 97, c
    if dpt == 'ZZ':
        return 0, c  # MP for french people living abroad


def dpt_to_int(dpt: str) -> int:
    """Convert a dpt to an int"""
    if type(dpt) is int:
        return int(dpt)

    if dpt[0] == 'Z':
        return 0  # Handling has to be done higher

    if dpt[1].upper() == 'A' or dpt[1].upper() == 'B':
        return 20

    else:
        return int(dpt)


def import_districts() -> Iterable[District]:
    """Import districts from insee_code_to_district.xlsx"""
    result = []
    wb = load_workbook("insee_code_to_district.xlsx")

    for row in wb["PR17_Découpage"].iter_rows(min_row=2):
        result.append(District(*(cell.value for cell in row)))

    logger.info("Imported %d districts.", len(result))

    return result


def import_insee_postal() -> Dict[str, str]:
    """Map INSEE code and postal codes from insee_postal.csv"""
    result = {}

    with open("insee_postal.csv") as f:
        reader = csv.reader(f, delimiter=';')
        for i, row in enumerate(reader):
            if i == 0:
                continue
            result[row[2]] = row[0]

    logger.info("Imported %d insee codes.", len(result))

    return result


def build_districts(districts: Iterable[District]) -> Dict[str, List[District]]:
    """Build relation between postal codes and districts numbers"""
    result = defaultdict(list)

    for d in districts:
        result[d.code_insee].append(d)

    return result


def bind(
    districts: Dict[str, List[District]], insee: Dict[str, str]
) -> Dict[str, Tuple[CodedDistrict]]:
    """Bind the two dicts together."""
    missing_districts = set()
    found: DefaultDict[str, Set[CodedDistrict]] = defaultdict(set)

    add_static_part(found)

    for cp, ci in insee.items():
        if ci in districts:
            found[cp].update(d.key for d in districts[ci])
        else:
            if cp not in found:
                missing_districts.add(ci)

    logger.info("Found %s districts over %s postal codes.", len(found), len(insee))
    logger.debug("Missing districts : %r", sorted(missing_districts))

    result = {key: tuple(value) for key, value in found.items()}

    for cp, ds in found.items():
        if len(ds) > 1:
            logger.debug("Found multiple circos for cp %s : %r", cp, ds)

    return result


def save(obj: object, filename):
    """Save the object into filename"""
    with open(filename, "w") as f:
        f.write(json.dumps(obj))
        logger.info("Object saved into %s", filename)


def main():
    """Main function."""
    logging.basicConfig(level="DEBUG")

    raw_districts = import_districts()
    insee_code_to_districts = build_districts(raw_districts)

    insee_postal = import_insee_postal()

    postal_code_to_district = bind(insee_code_to_districts, insee_postal)
    save(postal_code_to_district, "postal_code_to_district.json")


if __name__ == '__main__':
    main()
