from __future__ import absolute_import

import csv
import datetime as dt
from enum import Enum
import json
import io

import arrow
import openpyxl
import pytest
import xlrd
import xlsxwriter
from markupsafe import Markup
from pyquery import PyQuery

from webgrid import (
    BoolColumn,
    Column,
    DateTimeColumn,
    LinkColumnBase,
    NumericColumn,
    YesNoColumn,
    col_filter,
    col_styler,
    row_styler,
)
from webgrid.filters import TextFilter, OptionsEnumFilter
from webgrid.renderers import (
    CSV,
    HTML,
    XLS,
    XLSX,
    RenderLimitExceeded,
    render_html_attributes,
)
from webgrid_ta.grids import (
    ArrowCSVGrid,
    ArrowGrid,
    Grid,
    PeopleGrid as PG,
    StopwatchGrid,
)
from webgrid_ta.model.entities import (
    AccountType,
    ArrowRecord,
    Email,
    Person,
    Status,
    db,
)

from .helpers import eq_html, inrequest, render_in_grid


def _query_exclude_person(query):
    # this is pretty limited, but only used in the below couple of grids to
    # exclude the third Person record
    persons = Person.query.order_by(Person.id).limit(3).all()
    exclude_id = persons[2].id if len(persons) >= 3 else -1
    return query.filter(Person.id != exclude_id)


class PeopleGrid(PG):
    def query_prep(self, query, has_sort, has_filters):
        query = PG.query_prep(self, query, True, True)

        # default sort
        if not has_sort:
            query = query.order_by(Person.id.desc())

        # default filter
        if not has_filters:
            query = _query_exclude_person(query)

        return query


class PeopleCSVGrid(PG):
    allowed_export_targets = {'csv': CSV}

    def query_prep(self, query, has_sort, has_filters):
        query = PG.query_prep(self, query, True, True)

        # default sort
        if not has_sort:
            query = query.order_by(Person.id.desc())

        # default filter
        if not has_filters:
            query = _query_exclude_person(query)

        return query


def setup_module():
    Status.delete_cascaded()
    sp = Status(label='pending')
    sip = Status(label='in process')
    sc = Status(label='complete', flag_closed=1)
    db.session.add_all([sp, sip, sc])

    for x in range(1, 5):
        p = Person()
        p.firstname = 'fn%03d' % x
        p.lastname = 'ln%03d' % x
        p.sortorder = x
        p.numericcol = '2.13'
        p.state = 'st%03d' % x
        if x != 2:
            p.createdts = dt.datetime(2012, 0o2, 22, 10, x, 16)
            p.due_date = dt.date(2012, 0o2, x)
        db.session.add(p)
        p.emails.append(Email(email='email%03d@example.com' % x))
        p.emails.append(Email(email='email%03d@gmail.com' % x))
        if x % 4 == 1:
            p.status = sip
            p.account_type = AccountType.admin
        elif x % 4 == 2:
            p.status = sp
            p.account_type = AccountType.employee
        elif x % 4 == 0:
            p.status = None

    db.session.commit()


class SimpleGrid(Grid):
    on_page = 1
    per_page = 1

    Column('ID', 'id')
    Column('Name', 'name', filter=TextFilter(Person.firstname))
    Column('Status', 'status')
    Column('Emails', 'emails', can_sort=False)


class ColorColumn(Column):
    def format_data(self, data):
        if data == 'blue':
            return 'blue :)'
        return data


class EditColumn(LinkColumnBase):
    link_attrs = {'target': '_blank'}

    def create_url(self, record):
        return '/vehicle-edit/{0}'.format(record['id'])


class DealerColumn(LinkColumnBase):
    def create_url(self, record):
        return '/dealer-edit/{0}'.format(record['dealer_id'])

    def extract_data(self, record):
        return record['dealer'] + record['dealer_id']


class CarGrid(Grid):
    EditColumn('ID', 'id')
    EditColumn('Edit', 'edit', link_label='edit')
    DealerColumn('Dealer', 'dealer')
    Column('Make', 'make')
    Column('Model', 'model', class_='model')
    ColorColumn('Color', 'color')
    BoolColumn('Active', 'active')
    BoolColumn('Active Reverse', 'active', reverse=True)
    YesNoColumn('Active Yes/No', 'active')

    @row_styler
    def style_row(self, rownum, attrs, record):
        attrs.id = 'tr_{0}'.format(record['id'])

    @col_styler('model')
    def highlight_1500(self, attrs, record):
        if record['model'] == '1500':
            attrs.class_ += 'red'

    @col_filter('color')
    def pink_is_ugly(self, value):
        if value == 'pink':
            return 'pink :('
        return value


def find_tag(html, tag, id_=None, class_=None, **attrs):
    selector = tag
    if id_:
        selector += '#{}'.format(id_)
    if class_:
        selector += '.{}'.format(class_)
    for k, v in attrs.items():
        if v is None:
            selector += '[{}]'.format(k)
        else:
            selector += '[{}="{}"]'.format(k, v)

    return PyQuery(html)(selector)


def assert_tag(html, tag, text=None, **kwargs):
    results = find_tag(html, tag, **kwargs)
    assert results

    if text is not None:
        assert (
            any(i.text(squash_space=False) == text for i in results.items())
        ), '{} not found in {}'.format(text, results)
    return results


class TestHtmlRenderer(object):
    key_data = (
        {'id': 1, 'name': 'one'},
        {'id': 2, 'name': 'two'},
        {'id': 3, 'name': 'three'},
        {'id': 4, 'name': 'three'},
        {'id': 5, 'name': 'three'},
    )

    @inrequest('/')
    def test_car_html(self):
        key_data = (
            {'id': 1, 'make': 'ford', 'model': 'F150&', 'color': 'pink',
             'dealer': 'bob', 'dealer_id': '7', 'active': True,
             'active_1': True, 'active_2': True},
            {'id': 2, 'make': 'chevy', 'model': '1500', 'color': 'blue',
             'dealer': 'fred', 'dealer_id': '9', 'active': False,
             'active_1': False, 'active_2': False},
        )

        mg = CarGrid()
        mg.set_records(key_data)
        eq_html(mg.html.table(), 'basic_table.html')

    @pytest.mark.skipif(db.engine.dialect.name != 'sqlite', reason="IDs will not line up")
    @inrequest('/')
    def test_people_html(self):
        pg = render_in_grid(PeopleGrid, 'html')()
        eq_html(pg.html.table(), 'people_table.html')

    @pytest.mark.skipif(db.engine.dialect.name != 'sqlite', reason="IDs will not line up")
    @inrequest('/')
    def test_stopwatch_html(self):
        # Test Stopwatch grid with column groups.
        grid = StopwatchGrid()
        eq_html(grid.html.table(), 'stopwatch_table.html')

    @inrequest('/')
    def test_default_jinja_env(self):
        class TGrid(Grid):
            manager = None
            hide_controls_box = True
            session_on = False
            allowed_export_targets = None

            Column('ID', 'id', can_sort=False)
            Column('Value', 'value', can_sort=False)

        tg = TGrid()
        tg.set_records([
            {'id': 1, 'value': 'foo'},
        ])
        tg.html()

    def test_render_html_attributes(self):
        result = render_html_attributes({})
        assert isinstance(result, Markup)
        assert result == ''

        result = render_html_attributes({
            'text': 'abc',
            'empty': '',
            'bool1': True,
            'bool2': False,
            'none': None,
            'esc&': '<>"'
        })
        assert isinstance(result, Markup)
        assert result == ' bool1 empty="" esc&amp;="&lt;&gt;&#34;" text="abc"'

    @inrequest('/')
    def test_no_filters(self):
        class TGrid(Grid):
            Column('Test', Person.id)

        tg = TGrid()
        assert 'Add Filter' not in tg.html()

    @inrequest('/')
    def test_hide_excel_deprecated(self):
        class TGrid(Grid):
            hide_excel_link = True
            Column('Test', Person.id)
        with pytest.warns(
            DeprecationWarning,
            match='Hide excel link is deprecated, you should just override allowed_export_targets instead'  # noqa
        ):
            TGrid()

    def get_grid(self, **kwargs):
        g = SimpleGrid(**kwargs)
        g.set_records(self.key_data)
        g.apply_qs_args()
        return g

    @inrequest('/thepage?perpage=5&onpage=1')
    def test_current_url(self):
        g = self.get_grid()
        assert '/thepage?onpage=1&perpage=5' == g.html.current_url()
        assert '/thepage?onpage=1&perpage=10' == g.html.current_url(perpage=10)

    @inrequest('/thepage')
    def test_current_url_qs_prefix(self):
        g = self.get_grid(qs_prefix='dg_')
        assert '/thepage?dg_perpage=10' == g.html.current_url(perpage=10)

    @inrequest('/thepage?perpage=5&onpage=1&dgreset=1')
    def test_current_url_reset_removed(self):
        g = self.get_grid()
        assert '/thepage?onpage=1&perpage=5' == g.html.current_url()
        assert '/thepage?onpage=1&perpage=10' == g.html.current_url(perpage=10)

    @inrequest('/thepage?foo_dgreset=1')
    def test_current_url_reset_removed_prefix(self):
        g = self.get_grid(qs_prefix='foo_')
        assert '/thepage?foo_perpage=5' == g.html.current_url(perpage=5)

    @inrequest('/thepage?perpage=5&onpage=1')
    def test_current_url_reset_added(self):
        g = self.get_grid()
        assert '/thepage?dgreset=1&onpage=1&perpage=5' == g.html.current_url(dgreset=1)

    @inrequest('/thepage?perpage=5&onpage=1')
    def test_xls_url(self):
        g = self.get_grid()
        with pytest.warns(DeprecationWarning,
                          match='xls_url is deprecated. Use export_url instead.'):
            url = g.html.xls_url()
        assert url == '/thepage?export_to=xls&onpage=1&perpage=5'

    @inrequest('/thepage?perpage=5&onpage=1')
    def test_export_url(self):
        g = self.get_grid()
        assert g.html.export_url('xlsx') == '/thepage?export_to=xlsx&onpage=1&perpage=5'
        assert g.html.export_url('xls') == '/thepage?export_to=xls&onpage=1&perpage=5'
        assert g.html.export_url('csv') == '/thepage?export_to=csv&onpage=1&perpage=5'

    @inrequest('/thepage?onpage=3')
    def test_paging_url_first(self):
        g = self.get_grid()
        assert g.html.paging_url_first() == '/thepage?onpage=1&perpage=1'

    @inrequest('/thepage?onpage=3')
    def test_paging_url_next(self):
        g = self.get_grid()
        assert g.html.paging_url_next() == '/thepage?onpage=4&perpage=1'

    @inrequest('/thepage?onpage=3')
    def test_paging_url_prev(self):
        g = self.get_grid()
        assert g.html.paging_url_prev() == '/thepage?onpage=2&perpage=1'

    @inrequest('/thepage?onpage=3')
    def test_paging_url_last(self):
        g = self.get_grid()
        assert g.html.paging_url_last() == '/thepage?onpage=5&perpage=1'

    @inrequest('/thepage?foo=bar&onpage=5&perpage=10&sort1=1&sort2=2&sort3=3&op(name)=eq&v1(name)'
               '=bob&v2(name)=fred')
    def test_reset_url(self):
        g = self.get_grid()
        assert (
            g.html.reset_url()
            == '/thepage?dgreset=1&foo=bar&session_key={0}'.format(g.session_key)
        )

    @inrequest('/thepage?foo=bar&onpage=5')
    def test_form_action_url(self):
        g = self.get_grid()
        assert (
            g.html.form_action_url()
            == '/thepage?foo=bar&session_key={0}'.format(g.session_key)
        )

    @pytest.mark.parametrize('page_param,input_value', [
        (1, '1'),
        (2, '2'),
        (3, '3'),
        (4, '4'),
        (5, '5'),
        (0, '1'),
        (6, '5'),
        ('abc', '1'),
    ])
    def test_paging_select(self, page_param, input_value):

        @inrequest('/thepage?onpage={}'.format(page_param))
        def check_paging():
            g = self.get_grid()
            select_html = g.html.paging_select()
            assert PyQuery(select_html).text().strip() == 'of 5'
            assert_tag(select_html, 'input', id_='onpage', name='onpage', type='number',
                       value=input_value, min='1', max='5')

        check_paging()

    @inrequest('/thepage?onpage=2')
    def test_paging_html(self):
        g = self.get_grid()
        input_html = g.html.paging_input()
        assert_tag(input_html, 'input', name='perpage', type='number', value='1')

        img_html = g.html.paging_img_first()
        assert_tag(img_html, 'img', alt='<<', width='16', height='13',
                   src='/static/webgrid/b_firstpage.png')

        img_html = g.html.paging_img_first_dead()
        assert_tag(img_html, 'img', alt='<<', width='16', height='13',
                   src='/static/webgrid/bd_firstpage.png')

        img_html = g.html.paging_img_prev()
        assert_tag(img_html, 'img', alt='<', width='8', height='13',
                   src='/static/webgrid/b_prevpage.png')

        img_html = g.html.paging_img_prev_dead()
        assert_tag(img_html, 'img', alt='<', width='8', height='13',
                   src='/static/webgrid/bd_prevpage.png')

        img_html = g.html.paging_img_next()
        assert_tag(img_html, 'img', alt='>', width='8', height='13',
                   src='/static/webgrid/b_nextpage.png')

        img_html = g.html.paging_img_next_dead()
        assert_tag(img_html, 'img', alt='>', width='8', height='13',
                   src='/static/webgrid/bd_nextpage.png')

        img_html = g.html.paging_img_last()
        assert_tag(img_html, 'img', alt='>>', width='16', height='13',
                   src='/static/webgrid/b_lastpage.png')

        img_html = g.html.paging_img_last_dead()
        assert_tag(img_html, 'img', alt='>>', width='16', height='13',
                   src='/static/webgrid/bd_lastpage.png')

        # since we are on page 2, all links should be live
        footer_html = g.html.footer()
        assert g.html.paging_img_first() in footer_html
        assert g.html.paging_img_next() in footer_html
        assert g.html.paging_img_prev() in footer_html
        assert g.html.paging_img_last() in footer_html

        g.set_paging(1, 1)
        g.set_records(self.key_data)
        footer_html = g.html.footer()
        assert g.html.paging_img_first() not in footer_html, footer_html
        assert g.html.paging_img_first_dead() in footer_html
        assert g.html.paging_img_prev_dead() in footer_html

        g.set_paging(2, 3)
        g.set_records(self.key_data)
        footer_html = g.html.footer()
        assert g.html.paging_img_last() not in footer_html, footer_html
        assert g.html.paging_img_next_dead() in footer_html
        assert g.html.paging_img_last_dead() in footer_html

    @inrequest('/thepage?sort1=name&sort2=-id')
    def test_sorting_html(self):
        g = self.get_grid()

        select_html = g.html.sorting_select1()
        assert_tag(select_html, 'select', id_='sort1', name='sort1')
        assert_tag(select_html, 'option', value='', text='\N{NO-BREAK SPACE}')
        assert_tag(select_html, 'option', text='Name', selected=None, value='name')
        assert_tag(select_html, 'option', text='Name DESC', value='-name')
        assert_tag(select_html, 'option', text='ID', value='id')
        assert not find_tag(select_html, 'option', value='emails')

        select_html = g.html.sorting_select2()
        assert find_tag(select_html, 'option', text='Name', value='name').attr('selected') is None
        assert_tag(select_html, 'option', text='ID DESC', selected=None, value='-id')

        select_html = g.html.sorting_select3()
        assert_tag(select_html, 'option', selected=None, value='', text='\N{NO-BREAK SPACE}')

        heading_row = g.html.table_column_headings()
        assert 'sort-asc' not in heading_row
        assert 'sort-desc' not in heading_row

    @inrequest('/thepage?sort1=name')
    def test_sorting_headers_asc(self):
        g = self.get_grid()
        heading_row = g.html.table_column_headings()
        assert_tag(heading_row, 'a', text='Name', class_='sort-asc', href='/thepage?sort1=-name')

    @inrequest('/thepage?sort1=-name')
    def test_sorting_headers_desc(self):
        g = self.get_grid()
        heading_row = g.html.table_column_headings()
        assert_tag(heading_row, 'a', text='Name', class_='sort-desc', href='/thepage?sort1=name')

    @inrequest('/thepage?op(firstname)=eq&v1(firstname)=foo&op(createdts)=between&v1(createdts)='
               '2%2F15%2F12&&v2(createdts)=2012-02-16')
    def test_filtering_input_html(self):
        g = PeopleGrid()

        filter_html = g.html.filtering_col_inputs1(g.key_column_map['firstname'])
        assert '<input id="firstname_input1" name="v1(firstname)" type="text" />' in filter_html, \
            filter_html

        filter_html = g.html.filtering_col_inputs1(g.key_column_map['createdts'])
        assert '<input id="createdts_input1" name="v1(createdts)" type="text" />' in filter_html, \
            filter_html

        filter_html = g.html.filtering_col_inputs2(g.key_column_map['createdts'])
        assert '<input id="createdts_input2" name="v2(createdts)" type="text" />' in filter_html, \
            filter_html

        g.apply_qs_args()

        filter_html = g.html.filtering_col_inputs1(g.key_column_map['firstname'])
        assert '<input id="firstname_input1" name="v1(firstname)" type="text" value="foo" />' in \
            filter_html, filter_html

        filter_html = g.html.filtering_col_inputs1(g.key_column_map['createdts'])
        assert '<input id="createdts_input1" name="v1(createdts)" type="text" value=' + \
            '"02/15/2012 12:00 AM" />' in filter_html, filter_html

        filter_html = g.html.filtering_col_inputs2(g.key_column_map['createdts'])
        assert '<input id="createdts_input2" name="v2(createdts)" type="text" value=' + \
            '"02/16/2012 11:59 PM" />' in filter_html, filter_html

    @inrequest('/thepage?op(firstname)=foobar&v1(firstname)=baz')
    def test_filtering_invalid_operator(self):
        g = PeopleGrid()

        filter_html = g.html.filtering_col_inputs1(g.key_column_map['firstname'])
        assert '<input id="firstname_input1" name="v1(firstname)" type="text" />' in filter_html, \
            filter_html

    @inrequest('/thepage')
    def test_extra_filter_attrs(self):
        g = PeopleGrid()
        g.key_column_map['firstname'].filter.html_extra = {'data-special-attr': 'foo'}
        filter_html = g.html.filtering_table_row(g.key_column_map['firstname'])
        assert '<tr class="firstname_filter" data-special-attr="foo">' in filter_html, filter_html

    @inrequest('/thepage')
    def test_filter_primary_op_specified(self):
        g = PeopleGrid()
        g.key_column_map['firstname'].filter.primary_op = '!eq'
        filter_html = g.html.filtering_table_row(g.key_column_map['firstname'])
        assert PyQuery(filter_html).find('option[value="!eq"]').attr('data-render') == 'primary'
        assert not PyQuery(filter_html).find('option[value="eq"]').attr('data-render')

    @inrequest('/thepage')
    def test_filter_primary_op_not_specified(self):
        g = PeopleGrid()
        g.key_column_map['firstname'].filter.primary_op = None
        filter_html = g.html.filtering_table_row(g.key_column_map['firstname'])
        assert PyQuery(filter_html).find('option[value="eq"]').attr('data-render') == 'primary'
        assert not PyQuery(filter_html).find('option[value="!eq"]').attr('data-render')

    @inrequest('/thepage')
    def test_multiselect_enum_options(self):
        class PeopleType(Enum):
            bob = 'Bob'
            charlie = 'Charlie'

        g = PeopleGrid()
        name_filter = OptionsEnumFilter(Person.firstname, enum_type=PeopleType).new_instance()
        name_filter.set('is', ['bob'])
        output = g.html.filtering_filter_options_multi(name_filter, 'foo')
        assert PyQuery(output).find('input[value="bob"]').attr('checked')
        assert not PyQuery(output).find('input[value="charlie"]').attr('checked')

    @inrequest('/thepage')
    def test_confirm_export(self):
        g = PeopleGrid()
        assert json.loads(g.html.confirm_export()) == {'confirm_export': False, 'record_count': 3}

        g.unconfirmed_export_limit = 2
        assert json.loads(g.html.confirm_export()) == {'confirm_export': True, 'record_count': 3}

        g.unconfirmed_export_limit = None
        assert json.loads(g.html.confirm_export()) == {'confirm_export': False, 'record_count': 3}

    @inrequest('/thepage')
    def test_grid_rendering(self):
        g = PeopleGrid()
        # really just making sure no exceptions come through at this point
        assert g.html()

    @inrequest('/thepage')
    def test_no_records(self):
        g = PeopleGrid()
        g.set_records([])
        g.html
        assert '<p class="no-records">No records to display</p>' in g.html()

    @inrequest('/thepage')
    def test_no_pager(self):
        class PgNP(PeopleGrid):
            pager_on = False

        g = PgNP()
        assert '<td class="page">' not in g.html()
        assert '<td class="perpage">' not in g.html()
        assert '<th class="page">' not in g.html()
        assert '<th class="perpage">' not in g.html()

    def test_can_render(self):
        assert PeopleGrid().html.can_render() is True

    def test_render_error(self):
        class Renderer(HTML):
            def can_render(self):
                return False

        class TestGrid(PeopleGrid):
            def set_renderers(self):
                super(TestGrid, self).set_renderers()
                self.html = Renderer(self)

        with pytest.raises(RenderLimitExceeded):
            TestGrid().html()

    @inrequest('/thepage?search=foo')
    def test_render_search_filter(self):
        g = PeopleGrid()
        g.enable_search = True

        filter_html = g.html.filtering_fields()
        tag = assert_tag(filter_html, 'input', id='search_input', name='search', type='text')
        assert tag.val() == ''

        g.apply_qs_args()
        filter_html = g.html.filtering_fields()
        assert_tag(filter_html, 'input', id='search_input', name='search', type='text', value='foo')

    @inrequest('/thepage?search=foo&dgreset=1')
    def test_render_search_filter_reset(self):
        g = PeopleGrid()
        g.enable_search = True

        filter_html = g.html.filtering_fields()
        tag = assert_tag(filter_html, 'input', id='search_input', name='search', type='text')
        assert tag.val() == ''

        g.apply_qs_args()
        filter_html = g.html.filtering_fields()
        assert_tag(filter_html, 'input', id='search_input', name='search', type='text', value='')

    def test_search_disabled(self):
        class PeopleGrid2(PeopleGrid):
            enable_search = False

        g = PeopleGrid2()

        filter_html = g.html.filtering_fields()
        assert '<input id="search_input"' not in filter_html

    def test_no_searchable_columns(self):
        class TGrid(Grid):
            enable_search = True
            Column('Test', Person.id)

        tg = TGrid()
        filter_html = tg.html.filtering_fields()
        assert '<input id="search_input"' not in filter_html


class PGPageTotals(PeopleGrid):
    subtotals = 'page'


class TestPageTotals(object):
    @inrequest('/')
    def test_people_html(self):
        g = PGPageTotals()
        html = g.html()
        elem = find_tag(html, 'td', class_='totals-label', colspan='7')
        assert len(elem) == 1
        assert elem.text() == 'Page Totals (3 records):'


class PGGrandTotals(PeopleGrid):
    subtotals = 'grand'


class TestGrandTotals(object):
    @inrequest('/')
    def test_people_html(self):
        g = PGGrandTotals()
        g.html
        assert '<td class="totals-label" colspan="7">Grand Totals (3 records):</td>' in g.html()
        assert '<td class="totals-label" colspan="7">Page Totals (3 records):</td>' not in g.html()


class TestFooterRendersCorrectly(object):
    @inrequest('/')
    def test_people_html_footer(self):
        g = PeopleGrid()
        g.html
        assert '<a class="export-link" href="/?export_to=xlsx">XLSX</a>' in g.html()
        assert '<a class="export-link" href="/?export_to=xls">XLS</a>' in g.html()
        # make sure we are rendering the separator
        assert '&nbsp;|' in g.html()

    @inrequest('/')
    def test_people_html_footer_only_csv(self):
        g = PeopleCSVGrid()
        g.html
        assert '<a class="export-link" href="/?export_to=xls">XLS</a>' not in g.html()
        assert '<a class="export-link" href="/?export_to=csv">CSV</a>' in g.html()


class PGAllTotals(PeopleGrid):
    subtotals = 'all'


class TestAllTotals(object):
    @inrequest('/')
    def test_people_html(self):
        g = PGAllTotals()
        html = g.html()
        assert_tag(html, 'td', text='Grand Totals (3 records):', class_='totals-label', colspan='7')
        assert_tag(html, 'td', text='Page Totals (3 records):', class_='totals-label', colspan='7')


class PGTotalsStringExpr(PeopleGrid):
    subtotals = 'all'
    Column('FloatCol', 'float_col', has_subtotal=True)

    def query_prep(self, query, has_sort, has_filters):
        query = super(PGTotalsStringExpr, self).query_prep(query, has_sort, has_filters)
        return query.add_columns(Person.floatcol.label('float_col'))


class TestStringExprTotals:
    @inrequest('/')
    def test_people_html(self):
        g = PGTotalsStringExpr()
        html = g.html()

        assert_tag(html, 'td', text='Grand Totals (3 records):', class_='totals-label', colspan='7')
        assert_tag(html, 'td', text='Page Totals (3 records):', class_='totals-label', colspan='7')


class TestXLSRenderer(object):

    def test_some_basics(self):
        g = render_in_grid(PeopleGrid, 'xls')(per_page=1)
        buffer = io.BytesIO()
        wb = g.xls()
        wb.save(buffer)
        buffer.seek(0)

        book = xlrd.open_workbook(file_contents=buffer.getvalue())
        sh = book.sheet_by_name('render_in_grid')
        # headers
        assert sh.ncols == 10
        assert sh.cell_value(0, 0) == 'First Name'
        assert sh.cell_value(0, 7) == 'Sort Order'

        # last data row
        assert sh.cell_value(3, 0) == 'fn001'
        assert sh.cell_value(3, 7) == 1
        assert sh.nrows == 4

    def test_subtotals_with_no_records(self):
        g = PGGrandTotals()
        g.column('firstname').filter.op = 'eq'
        g.column('firstname').filter.value1 = 'foobar'
        buffer = io.BytesIO()
        wb = g.xls()
        wb.save(buffer)
        buffer.seek(0)

    def test_long_grid_name(self):
        class PeopleGridWithAReallyReallyLongName(PeopleGrid):
            pass
        g = PeopleGridWithAReallyReallyLongName()
        buffer = io.BytesIO()
        wb = g.xls()
        wb.save(buffer)
        buffer.seek(0)

        book = xlrd.open_workbook(file_contents=buffer.getvalue())
        book.sheet_by_name('people_grid_with_a_really_r...')

    def test_can_render(self):
        class FakeCountsGrid(PeopleGrid):
            def __init__(self, record_count, col_count, has_subtotals):
                self._num_records = record_count
                self._col_count = col_count
                self.subtotals = 'all' if has_subtotals else 'none'
                super(FakeCountsGrid, self).__init__()

            @property
            def record_count(self):
                return self._num_records

            def iter_columns(self, render_type):
                for _ in range(self._col_count):
                    yield None

        assert FakeCountsGrid(65535, 256, False).xls.can_render() is True
        assert FakeCountsGrid(65536, 256, False).xls.can_render() is False
        assert FakeCountsGrid(65535, 256, True).xls.can_render() is False
        assert FakeCountsGrid(65534, 256, True).xls.can_render() is True
        assert FakeCountsGrid(65535, 257, False).xls.can_render() is False

    def test_render_error(self):
        class Renderer(XLS):
            def can_render(self):
                return False

        class TestGrid(PeopleGrid):
            def set_renderers(self):
                super(TestGrid, self).set_renderers()
                self.xls = Renderer(self)

        with pytest.raises(RenderLimitExceeded):
            TestGrid().xls()


class TestXLSXRenderer(object):

    def test_some_basics(self):
        g = render_in_grid(PeopleGrid, 'xlsx')(per_page=1)
        wb = g.xlsx()
        wb.filename.seek(0)

        book = openpyxl.load_workbook(wb.filename)
        sh = book['render_in_grid']

        # headers
        assert sh.max_column == 10
        assert sh.cell(1, 1).value == 'First Name'
        assert sh.cell(1, 8).value == 'State'

        # last data row
        assert sh.max_row == 4
        assert sh.cell(4, 1).value == 'fn001'
        assert sh.cell(4, 8).value == 'st001'

    def test_group_headings(self):
        grid = StopwatchGrid()
        wb = grid.xlsx()
        wb.filename.seek(0)

        book = openpyxl.load_workbook(wb.filename)
        sheet = book[book.sheetnames[0]]
        #   [ A | B | C | D | E | F | G | H | I ]
        # 1 [       | Lap 1 |   | Lap 2 | Lap 3 ]
        row_values = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
        assert row_values == [None, None, 'Lap 1', None, None, 'Lap 2', None, 'Lap 3', None]
        assert sheet.cell(2, 2).value == 'Label'
        assert sheet.cell(3, 2).value == 'Watch 1'
        assert [str(range_) for range_ in sheet.merged_cells.ranges] == [
            'A1:B1',
            'C1:D1',
            # E is a single cell
            'F1:G1',
            'H1:I1',
        ]
        assert sheet.max_column == 9

    def test_subtotals_with_no_records(self):
        g = PGGrandTotals()
        g.column('firstname').filter.op = 'eq'
        g.column('firstname').filter.value1 = 'foobar'
        wb = g.xlsx()
        wb.filename.seek(0)

    def test_long_grid_name(self):
        class PeopleGridWithAReallyReallyLongName(PeopleGrid):
            pass
        g = PeopleGridWithAReallyReallyLongName()
        wb = g.xlsx()
        wb.filename.seek(0)

        book = openpyxl.load_workbook(wb.filename)
        assert book['people_grid_with_a_really_r...']

    def test_totals(self):
        g = PeopleGrid()
        g.subtotals = 'grand'

        wb = g.xlsx()
        wb.filename.seek(0)

        book = openpyxl.load_workbook(wb.filename)
        sheet = book[book.sheetnames[0]]
        assert sheet.max_row == 5
        assert sheet.cell(5, 1).value == 'Totals (3 records):'
        assert sheet.cell(5, 9).value == 6.39
        assert [str(range_) for range_ in sheet.merged_cells.ranges] == ['A5:H5']

    def test_totals_no_merge(self):
        class TestGrid(Grid):
            subtotals = 'all'
            Column('First Name', Person.firstname)
            NumericColumn('Number', Person.numericcol, has_subtotal=True)

        g = TestGrid()
        wb = g.xlsx()
        wb.filename.seek(0)

        book = openpyxl.load_workbook(wb.filename)
        sheet = book[book.sheetnames[0]]

        assert sheet.max_row == 6
        assert sheet.cell(6, 1).value == 'Totals (4 records):'
        assert sheet.merged_cells.ranges == []

    def test_can_render(self):
        class FakeCountsGrid(PeopleGrid):
            def __init__(self, record_count, col_count, has_subtotals):
                self._num_records = record_count
                self._col_count = col_count
                self.subtotals = 'all' if has_subtotals else 'none'
                super(FakeCountsGrid, self).__init__()

            @property
            def record_count(self):
                return self._num_records

            def iter_columns(self, render_type):
                for _ in range(self._col_count):
                    yield None

        assert FakeCountsGrid(1048575, 16384, False).xlsx.can_render() is True
        assert FakeCountsGrid(1048576, 16384, False).xlsx.can_render() is False
        assert FakeCountsGrid(1048575, 16384, True).xlsx.can_render() is False
        assert FakeCountsGrid(1048574, 16384, True).xlsx.can_render() is True
        assert FakeCountsGrid(1048575, 16385, False).xlsx.can_render() is False

    def test_render_error(self):
        class Renderer(XLSX):
            def can_render(self):
                return False

        class TestGrid(PeopleGrid):
            def set_renderers(self):
                super(TestGrid, self).set_renderers()
                self.xlsx = Renderer(self)

        with pytest.raises(RenderLimitExceeded):
            TestGrid().xlsx()

    def test_xlsx_format_caching(self):
        grid = PeopleGrid()
        wb = xlsxwriter.Workbook()
        format1 = grid.xlsx.style_for_column(wb, grid.column('status'))
        format2 = grid.xlsx.style_for_column(wb, grid.column('state'))
        format3 = grid.xlsx.style_for_column(wb, grid.column('numericcol'))

        assert format1 is not None
        assert format2 is not None
        assert format3 is not None

        assert format1 is format2
        assert format3 is not format1

        grid = PeopleGrid()
        grid.column('status').xlsx_style = {'bold': True, 'border': 1}
        grid.column('state').xlsx_style = {'bold': True, 'border': 2}

        format1 = grid.xlsx.style_for_column(wb, grid.column('status'))
        format2 = grid.xlsx.style_for_column(wb, grid.column('state'))
        assert format1 is not format2

        grid = PeopleGrid()
        grid.column('status').xlsx_style = {'bold': True, 'border': 1}
        grid.column('state').xlsx_style = {'bold': True, 'border': 1}

        format1 = grid.xlsx.style_for_column(wb, grid.column('status'))
        format2 = grid.xlsx.style_for_column(wb, grid.column('state'))
        assert format1 is format2


class TestCSVRenderer(object):

    def test_some_basics(self):
        g = render_in_grid(PeopleCSVGrid, 'csv')(per_page=1)
        csv_data = g.csv.build_csv()
        csv_data.seek(0)
        byte_str = io.StringIO(csv_data.read().decode('utf-8'))
        reader = csv.reader(byte_str, delimiter=',', quotechar='"')
        data = []
        for row in reader:
            data.append(row)
        assert len(data[0]) == 9
        assert data[0][0] == 'First Name'
        assert data[0][2] == 'Active'
        assert data[1][0] == 'fn004'

    def test_it_renders_date_time_with_tz(self):
        ArrowRecord.query.delete()
        ArrowRecord.testing_create(
            created_utc=arrow.Arrow(2016, 8, 10, 1, 2, 3)
        )
        g = ArrowCSVGrid()
        g.allowed_export_targets = {'csv': CSV}
        csv_data = g.csv.build_csv()
        csv_data.seek(0)
        byte_str = io.StringIO(csv_data.read().decode('utf-8'))
        reader = csv.reader(byte_str, delimiter=',', quotechar='"')
        data = []
        for row in reader:
            data.append(row)
        assert data[0][0] == 'Created'
        assert data[1][0] == '2016-08-10 01:02:03+0000'

    def test_it_renders_date_time_with_custom_format(self):
        class CSVGrid(Grid):
            session_on = True
            allowed_export_targets = {'csv': CSV}
            DateTimeColumn('Created', ArrowRecord.created_utc, csv_format='%m/%d/%Y %I:%M %p')

        ArrowRecord.query.delete()
        ArrowRecord.testing_create(
            created_utc=arrow.Arrow(2016, 8, 10, 1, 2, 3)
        )
        g = CSVGrid()
        csv_data = g.csv.build_csv()
        csv_data.seek(0)
        byte_str = io.StringIO(csv_data.read().decode('utf-8'))
        reader = csv.reader(byte_str, delimiter=',', quotechar='"')
        data = []
        for row in reader:
            data.append(row)
        assert data[0][0] == 'Created'
        assert data[1][0] == '08/10/2016 01:02 AM'


class TestHideSection(object):
    @inrequest('/')
    def test_controlls_hidden(self):
        class NoControlBoxGrid(PG):
            hide_controls_box = True
        g = NoControlBoxGrid()
        assert '<tr class="status"' not in g.html()
        assert '<div class="footer">' not in g.html()


class TestArrowDate(object):
    @inrequest('/')
    def test_arrow_render_html(self):
        ArrowRecord.query.delete()
        ArrowRecord.testing_create(created_utc=arrow.Arrow(2016, 8, 10, 1, 2, 3))
        g = ArrowGrid()
        assert '<td>08/10/2016 01:02 AM</td>' in g.html(), g.html()

        g.column('created_utc').html_format = 'YYYY-MM-DD HH:mm:ss ZZ'
        assert '<td>2016-08-10 01:02:03 +00:00</td>' in g.html(), g.html()

    @inrequest('/')
    def test_arrow_timezone(self):
        # regardless of timezone given, ArrowType stored as UTC and will display that way
        ArrowRecord.query.delete()
        ArrowRecord.testing_create(created_utc=arrow.Arrow(2016, 8, 10, 1, 2, 3).to('US/Pacific'))
        g = ArrowGrid()
        assert '<td>08/10/2016 01:02 AM</td>' in g.html(), g.html()

        g.column('created_utc').html_format = 'YYYY-MM-DD HH:mm:ss ZZ'
        assert '<td>2016-08-10 01:02:03 +00:00</td>' in g.html(), g.html()

    def test_xls(self):
        ArrowRecord.query.delete()
        ArrowRecord.testing_create(created_utc=arrow.Arrow(2016, 8, 10, 1, 2, 3))
        g = ArrowGrid()
        buffer = io.BytesIO()
        wb = g.xls()
        wb.save(buffer)
        buffer.seek(0)

        book = xlrd.open_workbook(file_contents=buffer.getvalue())
        sh = book.sheet_by_name('arrow_grid')
        # headers
        assert sh.cell_value(0, 0) == 'Created'
        # data row
        assert (
            dt.datetime(*xlrd.xldate_as_tuple(sh.cell_value(1, 0), sh.book.datemode)[:6])
            == dt.datetime(2016, 8, 10, 1, 2, 3)
        )
