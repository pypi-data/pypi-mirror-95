#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Mar 19 09:58:27 2019

@author: Andreas Geiges
"""
import re
import os
import pandas as pd
import numpy as np
from tqdm import tqdm
import deprecated as dp

from . import config

from . import core
from . import util
from . import mapping as mapp


 
from . import greenhouse_gas_database as gh
GHG_data = gh.GreenhouseGasTable()
from .data_structures import Datatable,TableSet
from openpyxl import load_workbook

REQUIRED_SETUP_FIELDS = ['filePath',
                         'fileName',
                         'sheetName',
                         'timeIdxList',
                         'spaceIdxList']

SP_ARG = 3
TI_ARG = 2
DT_ARG = 4

from datatoolbox.tools.excel import ExcelReader

#%% Functions

from .tools import excel as _xls

excel_to_pandas_idx = _xls.excel_to_pandas_idx
getAllSheetNames   = _xls.getAllSheetNames


def read_MAGICC6_MATLAB_bulkout(pathName):
    
    temp_offset = 0.61
    df = pd.read_table(pathName, skiprows=23, header=0, delim_whitespace=True, index_col=0)
    df = df + temp_offset
    df.index = df.index.astype(int)
    return df
    #df.values[df.values<0] = np.nan
#    
#    meta = dict()
#    meta['entity'] = 'global_temp'
#    meta['unit']   = '°C'
#    return Datatable(df, meta=meta)

def read_MAGICC6_BINOUT(filePath):
    import pymagicc as pym
    reader = pym.io._BinaryOutReader(filePath)

    metaData, df = reader.read()
    
    data = df.pivot(index='region',columns='time', values='value')
    
    metaDict = dict()
    metaDict['entity'] = df.variable[0]
    metaDict['source'] = 'MAGICC6_calculation'
    metaDict['unit']   = None

    return Datatable(data, meta=metaDict)

def read_PRIMAP_csv(fileName):
    
    metaMapping = {
            'entity': 'SHEET_ENTITY',
            'unit':'SHEET_UNIT',
            'category':'SHEET_NAME_CATEGORY',
            'scenario': 'SHEET_SCENARIO',
            'model' : 'SHEET_SOURCE'
            }
    
    allDf = pd.read_csv(fileName, usecols=[0,1], index_col=0)
    #print(os.path.basename(fileName))

    firstDataRow = allDf.loc['SHEET_FIRSTDATAROW','Unnamed: 1']
    
    #bugfix of old wrong formated PRIMAP files
    try:
        int(firstDataRow)
    except:
        firstDataRow = pd.np.where(allDf.index == "Countries\Years")[0][0]+3
   
    firstMetaRow = pd.np.where(allDf.index == "&SHEET_SPECIFICATIONS")[0][0]+1
    
    metaPrimap = dict()
    for row in range(firstMetaRow,firstDataRow):
        key = allDf.index[row]
        value = allDf.iloc[row,0]
        if key =='/':
            break
        metaPrimap[key] = value
    
    data = pd.read_csv(fileName, header=firstDataRow-2, index_col=0)
    
    meta = dict()
    for metaKey in metaMapping:
        
        if isinstance(metaMapping[metaKey], list):
            value = '_'.join(metaPrimap[x] for x in metaMapping[metaKey])
        else:
            value = metaPrimap[metaMapping[metaKey]]
        meta[metaKey] = value
    
    
    
    table = Datatable(data, meta=meta)
    table = table.loc[:, util.yearsColumnsOnly(table)]
    table.columns = table.columns.astype(int)
    return table#, data

def read_PRIMAP_Excel(fileName, sheet_names= None, xlsFile=None):
    
    if isinstance(sheet_names, str):
        return _read_PRIMAP_Excel_single(fileName, sheet_names, xlsFile)
    
    if sheet_names is None:
        sheet_names = getAllSheetNames(fileName)
        
    out = TableSet()
    if xlsFile is None:
        xlsFile = pd.ExcelFile(fileName)
    for sheet_name in tqdm(sheet_names):
        table = _read_PRIMAP_Excel_single(fileName, sheet_name, xlsFile=xlsFile)
        out[sheet_name] = table
        
    return out
        
def _read_PRIMAP_Excel_single(fileName, sheet_name= 0, xlsFile=None):
    if xlsFile is None: 
        xlsFile = pd.ExcelFile(fileName)
    allDf = pd.read_excel(xlsFile, sheet_name = sheet_name, usecols=[0,1], index_col=0)
    #print(os.path.basename(fileName))

    firstDataRow = allDf.loc['SHEET_FIRSTDATAROW','Unnamed: 1']
    
    #bugfix of old wrong formated PRIMAP files
    try:
        int(firstDataRow)
    except:
        firstDataRow = pd.np.where(allDf.index == "Countries\Years")[0][0]+3
        
    #print(firstDataRow)
    setup = dict()
    setup['filePath']  = os.path.dirname(fileName) +'/'
    setup['fileName']  = os.path.basename(fileName)
    setup['sheetName'] = sheet_name
    setup['timeIdxList']  = ('B' + str(firstDataRow-1), 'XX'+ str(firstDataRow-1))
    setup['spaceIdxList'] = ('A'+ str(firstDataRow), 'A1000')
    #print(setup)
    ex = ExcelReader(setup, xlsFile=xlsFile)
    data = ex.gatherData().astype(float)
    #return data
    meta = {'source': '',
            'entity': allDf.loc['SHEET_ENTITY','Unnamed: 1'],
            'unit': allDf.loc['SHEET_UNIT','Unnamed: 1'],
            'category': allDf.loc['SHEET_NAME_CATEGORY','Unnamed: 1'],
            'scenario': allDf.loc['SHEET_SCENARIO','Unnamed: 1'] + '|' + allDf.loc['SHEET_SOURCE','Unnamed: 1']}
    REG_ton = re.compile('^[GM]t')
    xx = REG_ton.search(meta['unit'])
    
    if xx:
        meta['unit'] = meta['unit'].replace(xx.group(0), xx.group(0) + ' ')
    
    table = Datatable(data, meta=meta)
    try:
        table = table.loc[:,util.yearsColumnsOnly(table)]
        table.columns = table.columns.astype(int)
#        table = table.loc[:,util.yearsColumnsOnly(table)]
    except:
        print('warning: Columns could not be converted to int')
    return table

def read_MAGICC6_ScenFile(fileName, **kwargs):
    VALID_MASS_UNITS= {
            'Pt': 1e18,
            'Gt': 1e15,
            'Mt': 1e12,
            'kt': 1e9,
            't' : 1e6,
            'Pg': 1e15,
            'Tg': 1e12,
            'Gg': 1e9,
            'Mg': 1e6,
            'kg': 1e3,
            'g' : 1}
    fid = open(fileName,'r')
    nDataRows    = int(fid.readline().replace('/n',''))
    
    while True:
    #for i, line in enumerate(fid.readlines()):
        line = fid.readline()
        if line[:11] == '{0: >11}'.format('YEARS'):
            break
    #get first header line
    entities  = line.split()[1:]
    
    
    #reading units
    unitLine = fid.readline().split()[1:]
    #print(unitLine)
    
    # find correct component
    components = [GHG_data.findEntryIdx(entity) for entity in entities]

    units = [unit for unit in unitLine if unit[:2] in VALID_MASS_UNITS]        
    
    
    replacementDict = {'MtN2O-N' : 'Mt N'}

    units = [replacementDict.get(unit, unit) for unit in units]

    columns = [(x,y,z) for x,y,z in zip(entities, components, units)]
    entityFrame =  pd.DataFrame(columns=entities)
    entityFrame.columns = pd.MultiIndex.from_tuples(columns)
    
    entityFrame.columns.names=['NAME', 'COMP','UNIT']

    for i, line in enumerate(fid.readlines()):
        
        if i == nDataRows:
            break
        data = line.split()
        entityFrame.loc[int(data[0])] = np.asarray(data[1:])

    # TODO: CHange to a results list of datatables
    return entityFrame

def insertDataIntoExcelFile(fileName, 
                            overwrite = False, 
                            setupSheet = 'INPUT',
                            interpolate = False):
    ins = _xls.ExcelWriter(fileName=fileName, overwrite = overwrite, setupSheet = setupSheet, interpolate=interpolate)
    ins.insert_data()
    ins.close()
    return ins
    

    
getAllSheetNames = _xls.getAllSheetNames


     
#%%
if False:
    #%%
    ws_readValues = load_workbook(self.setup['fileName'], data_only=True)['Sheet6']
    timeIdxString  = 'B2:K2'
    spaceIdxString = 'A4:A8'
    dataIdxString  = 'population|all|PROJECTION_MED|UN_WPP2017'
    
    ws_readValues = load_workbook(self.setup['fileName'], data_only=True)['Sheet7']
    timeIdxString  = 'B1:O1'
    spaceIdxString = 'DEU'
    dataIdxString  = 'A2:A6'
    
    ws_readValues = load_workbook(self.setup['fileName'], data_only=True)['Sheet8']
    timeIdxString  = '2020'
    spaceIdxString = 'B1:O1'
    dataIdxString  = 'A2:A6'

#Sheet2	B2:E2	A3:A4	
    ws_readValues = load_workbook(self.setup['fileName'], data_only=True)['Sheet2']
    timeIdxString  = 'B2:E2'
    spaceIdxString = 'A3:A4'
    dataIdxString  = 'area_agriculture|agriculture|historic|WDI2018'
##Sheet2	B2:E2	A3:A4	
#    ws_readValues = load_workbook(self.setup['fileName'], data_only=True)['Sheet4']
#    timeIdxString  = 'A'
#    spaceIdxString = 'B2:D2'
#    dataIdxString  = 'population|all|PROJECTION_LOW|UN_WPP2017'

    ws_readValues = load_workbook(self.setup['fileName'], data_only=True)['Sheet4a']
    timeIdxString  = 'A'
    spaceIdxString = '2'
    dataIdxString  = 'population|all|PROJECTION_LOW|UN_WPP2017'

    
    args =  None, None, None, None, None
    tableIDs = list()
    for args in iterData(dataIdxString, ws_readValues, *args):
        tableIDs.append(args[4])
    tables = dt.getTables(tableIDs)
    wksSheet = ws_readValues
    args =  None, None, None, None, None
    for argsTime in iterTime(timeIdxString, wksSheet, *args):
        if argsTime[TI_ARG] is None:
            continue
        for argsSpace in iterSpace(spaceIdxString, wksSheet, *argsTime):
            if argsSpace[SP_ARG] is None:
                continue
            for argData in iterData(dataIdxString, wksSheet, *argsSpace):
                if argData[DT_ARG] is None:
                    continue
                print(argData)
                try:
                    value = tables[argData[DT_ARG]].loc[argData[SP_ARG],int(argData[TI_ARG])]
                    print('success')
                except:
                    pass
#%%


def PandasExcelWriter(fileName):
    return pd.ExcelWriter(fileName,
                        engine='xlsxwriter',
                        datetime_format='mmm d yyyy hh:mm:ss',
                        date_format='mmmm dd yyyy')


#%%


        
if __name__ == '__main__':
    import datatools as dt
#    setup = dict()
#    setup['filePath']  = '../../projects/NDC-analysis/03_Ready/'
#    setup['fileName']  = 'CAT_CountryAssessment_Argentina_201812.xlsx'
#    setup['sheetName'] = 'OutputWebsite'
#    setup['timeIdxList']  = ('L40', 'AZ40')
#    setup['spaceIdxList'] = ('I42', 'I42')
#    
#    ex = Extractor(setup)
#    
#    print(ex.timeIdxList)
#    timeSeries = ex.gatherData()
#    
#    ex.setup(**{'spaceIdxList': ('I42', 'I43')})
#    dataDf = ex.gatherData()
#
#    ex.setup(**{'timeIdxList': ('L40', 'L40')})
#    spatSeries = ex.gatherData()    
#
#    ex.setup(**{'spaceIdxList': ('I42', 'I42')})
#    value = ex.gatherData() 
#    
#    excel_to_pandas_idx('AB12')    
    
    #%%
    from shutil import copyfile
    copyfile('demo_empty.xlsx', "demo.xlsx")
    ins = ExcelWriter(fileName='demo.xlsx')
    tables = ins.insert_data()
    ins.close()
    #%%    
    reader = ExcelReader_New(fileName='demo_filled.xlsx')
    out = reader.read_data()
#    for setup in ins.getSetups():
#        dataTable = dt.getTable(setup['dataID'])
#        ins._writeData(setup, dataTable)